import codecs
from io import BytesIO
import io
import html2text
from fpdf import FPDF
from django.core.files import File
from reportlab.lib.styles import getSampleStyleSheet
from urllib import response
from datetime import date
from urllib.parse import urlparse
from reportlab.platypus.tables import Table, TableStyle
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch,cm
from reportlab.lib.pagesizes import A4
from reportlab.graphics.shapes import Line,LineShape
from urllib.parse import urlparse 
from functools import partial
import os
from reportlab.platypus import PageBreak,SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle,LongTable,BaseDocTemplate,Flowable, Frame, PageTemplate
import urllib.request
from django.forms import PasswordInput
import xlrd
import csv
import numpy as np
from openpyxl import load_workbook
import re
import shutil
from google.cloud import storage
from django.views.generic.edit import FormView
import time
import cv2 as cv
import requests
from django.conf import settings
from django.shortcuts import render , redirect , HttpResponseRedirect
from django.views import  View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http.response import HttpResponse, JsonResponse
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from .token import account_activation_token
#from moviepy.editor import VideoFileClip, concatenate_videoclips
import os
from google.cloud import storage
from natsort import natsorted
import json
from django.core import serializers
import time
import pandas as pd
import random
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.views.decorators.csrf import csrf_exempt
import datetime
from collections import Counter
import secrets
import string
import datetime
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
import base64
import six
import uuid
from .models import PlanTable, RegistrationTbl, TalenstsumoFolder,MemberMapping
from .forms import FileFieldForm
import mysql.connector
from mysql.connector import Error
from .videoprocessor import videoProcess
from .videoprocessor1 import videoProcess1

class indexView(View):
    return_url = None
    def get(self , request,Id=0):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        userdbId = request.session.get('dbId')
        if(userId!=None):
            data = []
            connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
            cursor = connection.cursor()
            cursor.execute("SELECT tbl_Name,form_Name,form_Question_type,auto_report FROM `forms_templates` where published_Status=1 and user_Id ="+str(userdbId)+" and form_Question_type!='story' and form_Question_type!='course' and form_Question_type!='learning_plan';")
            rowforms = cursor.fetchall()
            for form in rowforms:
                data.append({'tblName':form[0],'frmName':form[1],'frmReport':form[3],'frmType':form[2]})
            cursor.execute("select member_of from tbl_member_mapping where userid= "+str(userdbId)+";")
            uid = cursor.fetchall()
            if(uid !=[]):
                for u in uid:
                    cursor.execute("SELECT tbl_Name,form_Name,form_Question_type,auto_report FROM `forms_templates` where published_Status=1 and user_Id ="+str(u[0])+" and public = 1 and form_Question_type!='story' and form_Question_type!='course' and form_Question_type!='learning_plan';")
                    rowforms = cursor.fetchall()
                    for form in rowforms:
                        data.append({'tblName':form[0],'frmName':form[1],'frmReport':form[3],'frmType':form[2]})
            cursor.close()
            return render(request , 'index.html',{'forms':data})
        else:
            return redirect('login')
            
class smtpView(View):
    return_url = None
    def get(self , request,Id=''):
        userId = request.session.get('userId')
        dbuserId = request.session.get('dbId')
        userEmail = request.session.get('userEmail')
        userdbId = request.session.get('dbId')
        if(userId!=None):
            return render(request , 'smtpsettings.html')
        else:
            return redirect('login')


class textView(View):
    return_url = None
    def get(self , request,Id=''):
        pass
    #     connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    #     cursor = connection.cursor()
    #     userdbId = request.session.get('dbId')
    #     getsettings= "SELECT table_name FROM information_schema.tables WHERE table_schema = 'interview' and table_name like 'ans_%';"
    #     cursor.execute(getsettings)
    #     rows = cursor.fetchall()
    #     vrowId = ''
    #     for tb in rows:
    #         getcolumns = "SELECT `COLUMN_NAME` as colm FROM `INFORMATION_SCHEMA`.`COLUMNS` WHERE `TABLE_SCHEMA`='interview' AND `TABLE_NAME`='"+tb[0]+"' AND `COLUMN_NAME` like 'interv_%';"
    #         cursor.execute(getcolumns)
    #         colmns = list(cursor.fetchall())
    #         vrowId = ''
    #         if(len(colmns)!=0):
    #             vrowId = ''
    #             rowmapping = "SELECT rowId FROM `video_collob_mapping` where tblName='"+tb[0]+"'"
    #             cursor.execute(rowmapping)
    #             rowID = cursor.fetchone()
    #             videRows = []
    #             if(rowID != None):
    #                 rowID = ''.join(rowID)
    #             else:
    #                 insermapping = "INSERT INTO `video_collob_mapping`(`tblName`, `rowId`) VALUES ('"+tb[0]+"','0');"
    #                 cursor.execute(insermapping)
    #                 connection.commit()
    #                 rowID = 0
    #             gettingVideoRows = "SELECT * FROM "+tb[0]+" WHERE '"+tb[0]+"id' > '"+rowID+"';"
    #             cursor.execute(gettingVideoRows)
    #             videRows = list(cursor.fetchall())
    #             field_names = [i[0] for i in cursor.description]
    #             if(len(videRows)!=0): 
    #                 colmindexed = []
    #                 for i,f in enumerate(field_names):
    #                     if  f.startswith('interv_'):
    #                         colmindexed.append(i)
    #                 vrowId = ''
    #                 for j,vRows  in enumerate(videRows):
    #                     if(j!=0):  
    #                         vurls = []
    #                         vrowId = str(videRows[j][0]).strip()
    #                         videlist = []
    #                         for k in colmindexed:
    #                             if(len(str(videRows[j][k]).strip())!=0):
    #                                 a = urlparse(str(videRows[j][k]).strip())
    #                                 vurls.append(os.path.basename(a.path))  
    #                         if(vurls != None or len(vurls) !=0):
    #                             cvideo = 1
    #                             for m,u in enumerate(vurls):
    #                                 video = cv.VideoCapture('media/videos/'+u)
    #                                 if (video.isOpened() == False):
    #                                     print("Error reading video file") 
    #                                 frame_width = int(video.get(3))
    #                                 frame_height = int(video.get(4))
    #                                 size = (frame_width, frame_height)
    #                                 fileName = 'media/mergevideo/filename'+str(m)+'.mp4'
    #                                 videlist.append(fileName)
    #                                 result = cv.VideoWriter(fileName,
    #                                                             cv.VideoWriter_fourcc(*'mp4v'),
    #                                                             10, size)   
    #                                 while(True):
    #                                     ret, frame = video.read()
    #                                     if ret == True:
    #                                         result.write(frame)
    #                                     else:
    #                                         break
    #                                 video.release()
    #                                 result.release()   
    #                                 cvideo = cvideo+1
    #                             if (int(cvideo)-1) == len(vurls) and len(videlist) !=0:   
    #                                 videoName = "collobt_"+tb[0]+"_r_"+ vrowId + '.mp4'
    #                                 vurl = 'http://127.0.0.1:8000/media/videos/' + videoName                
    #                                 settingVideoRows = "UPDATE `"+tb[0]+"` SET `videoColloboration` = '"+vurl+"' WHERE `"+tb[0]+"id` = '"+vrowId+"';"
    #                                 cursor.execute(settingVideoRows)
    #                                 connection.commit()
    #                                 concatinated = concatenate_videoclips([VideoFileClip(v) for v in videlist])
    #                                 concatinated.write_videofile('media/videos/' + videoName, codec='libx264')
    #                                 shutil.rmtree('media/mergevideo/') 
    #                                 os.mkdir('media/mergevideo/')
    #             if(len(vrowId)!= 0):
    #                 updatetblMapping = "UPDATE `video_collob_mapping` SET `rowId`='"+vrowId+"' where `tblName` = '"+tb[0]+"';"
    #                 cursor.execute(updatetblMapping)
    #                 connection.commit()
    #     cursor.close()
    #     return render(request , 'test.html')


@csrf_exempt
def getSettingsData(request):
    if request.method == 'POST':
        data = []
        defaultval = '';
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        userdbId = request.session.get('dbId')
        getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid = '"+str(userdbId)+"';"
        cursor.execute(getsettings)
        results = cursor.fetchall()
        if(cursor.rowcount == 0):
            getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid  = 51;"
            cursor.execute(getsettings)
            results = cursor.fetchall()
            defaultval = "Default"
            cursor.close()
        response = {"data":results,"default":defaultval}
        return JsonResponse(response,safe=False)


@csrf_exempt
def get_listsData(request):
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    userdbId = request.session.get('dbId')
    getuserids= "SELECT * FROM `tbl_students_lists` WHERE `user_Id` ='"+str(userdbId)+"';"
    cursor.execute(getuserids)
    allrows = cursor.fetchall()
    cursor.close()
    return JsonResponse(allrows,safe=False)

def get_studentsData(request):
    if(request.method == 'GET'):
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        userdbId = request.session.get('dbId')
        listName = request.GET['ulistName']
        getuserids= "SELECT `tbl_students_details`.`id` as id,`student_name`, `student_email`, `student_collage`, `student_specialization`, `student_course`, `student_year`, `student_active` FROM `tbl_students_details` INNER JOIN `tbl_students_lists` ON `tbl_students_details`.student_list_id = `tbl_students_lists`.`id` WHERE `tbl_students_lists`.`user_id`='"+str(userdbId)+"' AND `tbl_students_lists`.id= '"+str(listName)+"' ORDER BY `tbl_students_lists`.`user_id`;"
        cursor.execute(getuserids)
        allrows = cursor.fetchall()
        cursor.close()
        return JsonResponse(allrows,safe=False)

             
def logout(request):
    try:
        del request.session["userId"]
        del request.session["userEmail"]
        del request.session["userName"]
        del request.session["dbId"]
        del request.session["Plan"]
        del request.session["Number"]
    except:
        print("error")
    return redirect('login')
    

class studentReviewView(View):
    return_url = None
    def get(self , request,Id,formType,tblName):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        if(userId!=None):
            rows = []
            columns = []
            final = []
            list1 = []
            list2 = []
            colms1 = []
            fields = []
            nomineelist = ['nominee','share','nominate','send']
            revielsts = ['feedback','reviewer','review']
            interviewlist = ['interview','link']
            interpattern = '|'.join(interviewlist)
            cursor = connection.cursor()
            cursor.callproc('getdetails1', [tblName, Id,])
            rowdata = ''
            rowdata1 = ''
            for result in cursor.stored_results():
                fields = [x[0] for x in result.description]
                rowdata1=result.fetchall()
            for prow in rowdata1:
                rows = list(prow)
            idcolum = tblName+'id'
            exclst =  ['ipAddress','SubDate','reviewStatus','sent','sentDate','commReview','overGrade','overGrade','sentReport','ReportLink','AutoGrade',idcolum]
            cursor.close()
            cursor1 = connection.cursor()
            cursor1.callproc('getcolmnData1', [tblName, ])
            coludata = ''
            coludata1 = ''
            for result in cursor1.stored_results():
                coludata1=result.fetchall()
            for pcol in coludata1:
                columns = list(pcol)
            df = pd.DataFrame({'ques':columns,'ans':rows,'fields':fields})
            final = []
            nomineeemail = ''
            revieweremail = ''
            getcolumns= "SELECT `name`, `email`, `review`,`nominee` FROM form_column_mapping WHERE tblName = '"+tblName.strip()+"';"
            cursor1.execute(getcolumns)
            results = cursor1.fetchall()
            for i, t in enumerate(results):
                if(i ==0):
                    nomineeemail = t[2] 
                    revieweremail = t[3]
            if len(nomineeemail) != 0:
                exclst.insert(len(exclst),nomineeemail)
            if len(revieweremail) != 0:
                exclst.insert(len(exclst),revieweremail)
            cursor1.close()
            interlist = df.index[df['fields'].str.contains('interv_')]
            ilist = interlist.tolist()
            rlist = []
            for i, j in enumerate(ilist):
                rlist.append(int(df[df["fields"]==df['fields'][j].replace("interv_","review_")].index.values))
            for ind in df.index:
                if(str(df['ans'][ind]).strip() == 'name' or str(df['ans'][ind]).strip() == 'email' or len(str(df['ans'][ind]).strip()) == 0):
                    df['ans'][ind] = 'N/A'
                else:
                    df['ans'][ind] = df['ans'][ind]
                    if (ind not in rlist and df['fields'][ind] not in exclst):
                        if(df['fields'][ind].startswith("lang_") == False):
                            final.append({"que":(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":"No","field":df['fields'][ind]})
            for k,ind in enumerate(interlist):
                if(str(df['ans'][rlist[k]]).strip() == 'None'):
                    df['ans'][rlist[k]] = 'N/A'
                final.append({"que":(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":df['ans'][rlist[k]],"field":df['fields'][ind]})
            common_df = df.loc[df['fields'] == 'commReview']
            if(str(common_df['ans'].values[0]).strip() == 'None'):
                common_df['ans'].values[0] = 'N/A'
            final.append({"que":(common_df['ques'].values[0]).replace("_"," "),"ans":common_df['ans'].values[0],"rev":"No","field":common_df['fields'].values[0]})
            if(formType == 'quiz'):
                auto_df = df.loc[df['fields'] == 'AutoGrade']
                final.append({"que":(auto_df['ques'].values[0]).replace("_"," "),"ans":auto_df['ans'].values[0],"rev":"No","field":auto_df['fields'].values[0]})
            grade_df = df.loc[df['fields'] == 'overGrade']
            final.append({"que":(grade_df['ques'].values[0]).replace("_"," "),"ans":grade_df['ans'].values[0],"rev":"No","field":grade_df['fields'].values[0]})
            return render(request , 'studentReview.html',{'data':final})


@csrf_exempt
def get_TeamData(request):
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    userdbId = request.session.get('dbId')
    getuserids= "SELECT userid FROM `tbl_member_mapping` WHERE member_of='"+str(userdbId)+"';"
    cursor.execute(getuserids)
    allrows = cursor.fetchall()
    uids = []
    for i in allrows:
        uidvalue = list(i)
        uids.append("'"+str(uidvalue[0])+"'")
    strnids=','.join([str(x) for x in uids])
    if(len(strnids.strip()) != 0):
        getcolumns= "SELECT * FROM `registration_tbl` WHERE id in ("+str(strnids)+");"
        cursor.execute(getcolumns)
        results = cursor.fetchall()
        cursor.close()
    else:
        results = []
    return JsonResponse(results,safe=False)

@csrf_exempt
def intgetformData(request):
    if request.method == 'POST':
        data = []
        uformId = request.POST.get("formId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        getcolumns= "SELECT `name`, `email`, `review`,`nominee` FROM form_column_mapping WHERE tblName = '"+uformId.strip()+"';"
        cursor.execute(getcolumns)
        results = cursor.fetchall()
        cursor.close()
        try:
            for i, t in enumerate(results):
                if(i ==0):
                    fircolm =t[0] 
                    seccolm =t[1]
                    thirdcolm =t[2] 
                    fourthcolm =t[3]
                    if((len(fircolm) !=0) and (len(seccolm) !=0)):
                        gettbldata= "SELECT `"+uformId.strip()+"id`,`"+fircolm.strip()+"`,`"+seccolm.strip()+"`,DATE_FORMAT(`SubDate`,'%d/%m/%Y') AS SubDate,`commReview`,`reviewStatus`,`overGrade`,`sent`,`sentReport`,`ReportLink` FROM "+uformId.strip()+";"
                        cursor1 = connection.cursor()
                        cursor1.execute(gettbldata)
                        tblresults = cursor1.fetchall()
                        cursor1.close()
                        for tblrows in tblresults:
                            data.append({'id':tblrows[0],'name':tblrows[1],'email':tblrows[2],'subon':tblrows[3],'commRev':tblrows[4],'revStatus':tblrows[5],'grade':tblrows[6],'repstatus':tblrows[7],'scorereport':tblrows[8],'scoreLink':tblrows[9]})
                        data.pop(0)
                    else:
                        gettbldata= "SELECT `"+uformId.strip()+"id`,DATE_FORMAT(`SubDate`,'%d/%m/%Y') AS SubDate,`commReview`,`reviewStatus`,`overGrade`,`sent`,`sentReport`,`ReportLink` FROM "+uformId.strip()+";"
                        cursor1 = connection.cursor()
                        cursor1.execute(gettbldata)
                        tblresults = cursor1.fetchall()
                        cursor1.close()
                        for tblrows in tblresults:
                            data.append({'id':tblrows[0],'name':'','email':'','subon':tblrows[1],'commRev':tblrows[2],'revStatus':tblrows[3],'grade':tblrows[4],'repstatus':tblrows[5],'scorereport':tblrows[6],'scoreLink':tblrows[7]})
                        data.pop(0)
        except:
            data = []
        return JsonResponse(data,safe=False)

        
class loginView(View):
    return_url = None
    def get(self , request,Id=0):
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        userdbId = request.session.get('dbId')
        if(userId!=None):
            data = []
            try:
                cursor = connection.cursor()
                cursor.execute("SELECT tbl_Name,form_Name,form_Question_type,auto_report FROM `forms_templates` where published_Status=1 and user_Id ="+str(userdbId)+" and form_Question_type!='story' and form_Question_type!='course' and form_Question_type!='learning_plan';")
                rowforms = cursor.fetchall()
                for form in rowforms:
                    data.appenddata.append({'tblName':form[0],'frmName':form[1],'frmReport':form[3],'frmType':form[2]})
                cursor.execute("select member_of from tbl_member_mapping where userid= "+str(userdbId)+";")
                uid = cursor.fetchall()
                if(uid !=[]):
                    for u in uid:
                        cursor.execute("SELECT tbl_Name,form_Name,form_Question_type,auto_report FROM `forms_templates` where published_Status=1 and user_Id ="+str(u[0])+" and public = 1 and form_Question_type!='story' and form_Question_type!='course' and form_Question_type!='learning_plan';")
                        rowforms = cursor.fetchall()
                        for form in rowforms:
                            data.appenddata.append({'tblName':form[0],'frmName':form[1],'frmReport':form[3],'frmType':form[2]})
                cursor.close()
                return render(request , 'index.html',{'forms':data})
            except:
                del request.session["userId"]
                del request.session["userEmail"]
                del request.session["userName"]
                del request.session["dbId"]
                del request.session["Plan"]
                del request.session["Number"]               
                return render(request, 'login.html')
            
        else:
            return render(request, 'login.html')

@csrf_exempt
def update_reviews_view(request):
    if request.method == 'POST':
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        utbl = request.POST.get("utbl")
        uid = request.POST.get("uparam")
        uname = request.POST.get("uname")
        uemail = request.POST.get("uemail")
        usubDate = request.POST.get("uSubDate")
        udata = request.POST['udata']
        data = json.loads(udata)
        updatrev =  'UPDATE '+utbl+' SET '
        for i, d in enumerate(data):
            name = d['fieldName']
            value = d['fieldValue']
            if "interv_" in name.lower():
                fname = name.replace('interv_',"review_")
                updatrev+= fname+" = '"+str(value) + "',"
            else:
                updatrev+= name+" = '"+str(value) + "',"
        today = datetime.datetime.now()
        date_time = today.strftime("%m/%d/%Y, %H:%M:%S")
        updatrev+='reviewStatus="Yes",sent="Yes", sentDate="'+ date_time +'"'
        updatrev+=' WHERE '+utbl+'id ='+uid+';'
        cursor = connection.cursor()
        cursor.execute(updatrev)
        connection.commit()
        cursor.close()
        link = 'http://test.techdivaa.com/review/'+uid+'/'+utbl+''
        #link = 'http://127.0.0.1:8000/review/'+uid+'/'+utbl+''
        cursor1 = connection.cursor()
        cursor1.execute("SELECT form_Name,form_ID FROM `forms_templates` WHERE tbl_Name="+'"'+utbl+'"' )
        rows = cursor1.fetchone()
        cursor1.close()
        sendEmailinterview(request.session.get('dbId'),'reviewlink',uname,uemail,rows[0],rows[1],usubDate,link,utbl,'','','')
        return HttpResponse('Updated')
	    
@csrf_exempt
def update_settings_view(request):
    if request.method == 'POST':
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        userId = request.session.get('dbId')
        udata = request.POST['udata']
        data = json.loads(udata)
        cursor = connection.cursor()
        stmt = "select * from  settings_tbl where userid = '"+str(userId)+"'"
        cursor.execute(stmt)
        result = cursor.fetchall()
        cursor.close()
        if(len(result) == 0):
            for i, d in enumerate(data):
                field =  d['settings_Field']
                fieldval = d['settings_Val']
                if('xx' in fieldval):
                    if(field == 'host'):
                        fieldval = 'smtp.acumbamail.com'
                    elif(field == 'login'):
                        fieldval = 'deb@talentsumo.co'
                    elif(field == 'password'):
                        fieldval = '9sHNPoWF7dXBWfrfdtmQ'
                    elif(field == 'from'):
                        fieldval = 'report@talentsumo.co'
                cursor1 = connection.cursor()
                stmt1 = "INSERT INTO `settings_tbl`( `settings_Field`, `settings_Value`, `userid`) VALUES ('"+field+"','"+fieldval+"','"+str(userId)+"');"
                cursor1.execute(stmt1)
                connection.commit()
                cursor1.close()
        else:
            for i, d in enumerate(data):
                field =  d['settings_Field']
                fieldval = d['settings_Val']
                if(field=="host" or field=="port" or field=="login" or field=="password" or field=="from" or field=="cEmailStatus" or field=="rEmailStatus" or field=="nEmailStatus" or field=="fEmailStatus" or field=="campEmailStatus" or field=="arptEmailStatus" or field=="arptEmailStatus" or field=="arptlogo" or field=="arpfooterText" or field=="PdfSignature" ):
                    if('xx' in fieldval):
                        if(field == 'host'):
                            fieldval = 'smtp.acumbamail.com'
                        elif(field == 'login'):
                            fieldval = 'deb@talentsumo.co'
                        elif(field == 'password'):
                            fieldval = '9sHNPoWF7dXBWfrfdtmQ'
                        elif(field == 'from'):
                            fieldval = 'report@talentsumo.co'
                    cursor1 = connection.cursor()
                    stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                    cursor1.execute(stmt1)
                    connection.commit()
                    cursor1.close()
                    if(field=="cEmailStatus"):
                        ctextdata = data[6]
                        ffield =  ctextdata['settings_Field']
                        ffieldval = ctextdata['settings_Val']
                        cursor1 = connection.cursor()
                        stmt1 ="UPDATE `settings_tbl` SET `settings_Value` = '"+ffieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+ffield+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        stmt2 ="UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt2)
                        connection.commit()
                        cursor1.close()
                    elif(field=="rEmailStatus"):
                        rtextdata = data[8]
                        ffield =  rtextdata['settings_Field']
                        ffieldval = rtextdata['settings_Val']
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+ffieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+ffield+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        stmt2 ="UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt2)
                        connection.commit()
                        cursor1.close()
                    elif(field=="nEmailStatus"):
                        ntextdata = data[10]
                        ffield =  ntextdata['settings_Field']
                        ffieldval = ntextdata['settings_Val']
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+ffieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+ffield+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        stmt2 ="UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt2)
                        connection.commit()
                        cursor1.close()
                    elif(field=="fEmailStatus"):
                        ftextdata = data[12]
                        ffield =  ftextdata['settings_Field']
                        ffieldval = ftextdata['settings_Val']
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+ffieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+ffield+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        stmt2 ="UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt2)
                        connection.commit()
                        cursor1.close()
                    elif(field=="campEmailStatus"):
                        ftextdata = data[14]
                        ffield =  ftextdata['settings_Field']
                        ffieldval = ftextdata['settings_Val']
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+ffieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+ffield+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        stmt2 ="UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt2)
                        connection.commit()
                        cursor1.close()
                    elif(field=="arptEmailStatus"):
                        ftextdata = data[16]
                        ffield =  ftextdata['settings_Field']
                        ffieldval = ftextdata['settings_Val']
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+ffieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+ffield+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        stmt2 ="UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt2)
                        connection.commit()
                        cursor1.close()
                    elif(field=="arptlogo"):
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        cursor1.close()
                    elif(field=="arpfooterText"):
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        cursor1.close()
                    elif(field=="PdfSignature"):
                        cursor1 = connection.cursor()
                        stmt1 = "UPDATE `settings_tbl` SET `settings_Value` = '"+fieldval+"'  WHERE userid = '"+str(userId)+"' and settings_Field = '"+field+"';"
                        cursor1.execute(stmt1)
                        connection.commit()
                        cursor1.close()
        return HttpResponse('Updated')

# Create your views here.
class reviewView(View):
    return_url = None
    def get(self , request,Id,formType,tblName):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        if(userId!=None):
            rows = []
            columns = []
            final = []
            list1 = []
            list2 = []
            colms1 = []
            fields = []
            nomineelist = ['nominee','share','nominate','send']
            revielsts = ['feedback','reviewer','review']
            interviewlist = ['interview','link']
            interpattern = '|'.join(interviewlist)
            cursor = connection.cursor()
            cursor.execute("SELECT  `videocolloboration` FROM `forms_templates` WHERE `tbl_Name` = '"+str(tblName)+"';")
            videocollActive = cursor.fetchone()
            cursor.callproc('getdetails1', [tblName, Id,])
            rowdata = ''
            rowdata1 = ''
            for result in cursor.stored_results():
                fields = [x[0] for x in result.description]
                rowdata1=result.fetchall()
            for prow in rowdata1:
                rows = list(prow)
            idcolum = tblName+'id'
            exclst =  ['ipAddress','SubDate','reviewStatus','sent','sentDate','commReview','overGrade','overGrade','sentReport','ReportLink','AutoGrade',idcolum]
            cursor.close()
            cursor1 = connection.cursor()
            cursor1.callproc('getcolmnData1', [tblName, ])
            coludata = ''
            coludata1 = ''
            for result in cursor1.stored_results():
                coludata1=result.fetchall()
            for pcol in coludata1:
                columns = list(pcol)
            df = pd.DataFrame({'ques':columns,'ans':rows,'fields':fields})
            final = []
            nomineeemail = ''
            revieweremail = ''
            getcolumns= "SELECT `name`, `email`, `review`,`nominee` FROM form_column_mapping WHERE tblName = '"+tblName.strip()+"';"
            cursor1.execute(getcolumns)
            results = cursor1.fetchall()
            for i, t in enumerate(results):
                if(i ==0):
                    nomineeemail = t[2] 
                    revieweremail = t[3]
            if len(nomineeemail) != 0:
                exclst.insert(len(exclst),nomineeemail)
            if len(revieweremail) != 0:
                exclst.insert(len(exclst),revieweremail)
            cursor1.close()
            interlist = df.index[df['fields'].str.contains('interv_')]
            ilist = interlist.tolist()
            rlist = []
            for i, j in enumerate(ilist):
                rlist.append(int(df[df["fields"]==df['fields'][j].replace("interv_","review_")].index.values))
            for ind in df.index:
                if(str(df['ans'][ind]).strip() == 'name' or str(df['ans'][ind]).strip() == 'email' or len(str(df['ans'][ind]).strip()) == 0):
                    df['ans'][ind] = 'N/A'
                else:
                    df['ans'][ind] = df['ans'][ind]
                    if (ind not in rlist and df['fields'][ind] not in exclst):
                        if(df['fields'][ind].startswith("lang_") == False):
                            final.append({"que":(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":"No","field":df['fields'][ind]})
            for k,ind in enumerate(interlist):
                if(str(df['ans'][rlist[k]]).strip() == 'None'):
                    df['ans'][rlist[k]] = 'N/A'
                final.append({"que":(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":df['ans'][rlist[k]],"field":df['fields'][ind]})
            common_df = df.loc[df['fields'] == 'commReview']
            if(str(common_df['ans'].values[0]).strip() == 'None'):
                common_df['ans'].values[0] = 'N/A'
            final.append({"que":(common_df['ques'].values[0]).replace("_"," "),"ans":common_df['ans'].values[0],"rev":"No","field":common_df['fields'].values[0]})
            if(formType == 'quiz'):
                auto_df = df.loc[df['fields'] == 'AutoGrade']
                final.append({"que":(auto_df['ques'].values[0]).replace("_"," "),"ans":auto_df['ans'].values[0],"rev":"No","field":auto_df['fields'].values[0]})
            grade_df = df.loc[df['fields'] == 'overGrade']
            final.append({"que":(grade_df['ques'].values[0]).replace("_"," "),"ans":grade_df['ans'].values[0],"rev":"No","field":grade_df['fields'].values[0]})
            return render(request , 'review.html',{'data':final,'video':int(videocollActive[0])})
        else:
            return redirect('login')

def generateTableSchema(tableNm,prmid,colmnlist):
    schematb = 'CREATE TABLE '
    schematb+= tableNm
    schematb+='('+prmid+' INTEGER '
    for c in colmnlist:
        schematb+=','+c+' TEXT NULL'
    schematb+=');'
    return schematb


def generateInsertSchema(tableNm,colmnlist,row2):
    if(len(row2)<len(colmnlist)):
        for i in  range(len(row2),len(colmnlist)):
            row2.insert(i, "")
    collist = ",".join(colmnlist)
    answsist = ', '.join(['"{}"'.format(value) for value in row2])
    insertschema = 'INSERT INTO '+tableNm+'('
    insertschema+=collist
    insertschema+=')VALUES('
    insertschema+=answsist
    insertschema+=');'
    return insertschema
    
def getcolumns(tableName):
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    getcolumns= "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '"+tableName+"';"
    cursor.execute(getcolumns)
    results = cursor.fetchall()
    dbcolumnlist =[]
    for t in results:
        dbcolumnlist.append(''.join(t))
    cursor.close()
    return dbcolumnlist

def sendEmailinterview(userdbId,type,name,email,formname,formId,submittedAt,revlink,utbl_Name,answerlink,noemail,reveiweremail):
    inteviewlink = 'https://spellcheck.techdivaa.com/'+formId
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid = '"+str(userdbId)+"';"
    cursor.execute(getsettings)
    results = cursor.fetchall()
    if(cursor.rowcount == 0):
        getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid  = 51;"
        cursor.execute(getsettings)
        results = cursor.fetchall()
    cursor.close()
    host = ''
    port = ''
    login = ''  
    password = ''
    fromaddr = '' 
    message = ''
    if(len(results) != 0):
        host = results[0][1]
        port = results[1][1]
        login = results[2][1]
        password = results[3][1]
        fromaddr =  results[4][1]
        if(type=="reviewlink"):
            rlnkstatus = results[11][1]
            if(rlnkstatus == 'true'):
                rlnktext = results[12][1]
                rlnktext = str(rlnktext)
                rlnktext = rlnktext.replace('@Name', name)
                rlnktext = rlnktext.replace('@Email', email)
                rlnktext = rlnktext.replace('@FormName', formname)
                rlnktext = rlnktext.replace('@SubmittedAt', submittedAt)
                rlnktext = rlnktext.replace('@InterviewLink', answerlink)
                rlnktext = rlnktext.replace('@FormLink', inteviewlink)
                rlnktext = rlnktext.replace("@ReviewFormLink", revlink)
                mailstatus = SendEmailAct(fromaddr,email,login,password,host,port,rlnktext,'Review Link')
        elif(type=="candidate"):
            canstatus = results[5][1]
            if(canstatus == 'true'):
                cantext = results[6][1]
                cantext = str(cantext)
                cantext = cantext.replace('@Name', name)
                cantext = cantext.replace('@Email', email)
                cantext = cantext.replace('@FormName', formname)
                cantext = cantext.replace('@SubmittedAt', submittedAt)
                cantext = cantext.replace('@InterviewLink', answerlink)
                cantext = cantext.replace('@FormLink', inteviewlink)
                cantext = cantext.replace("@ReviewFormLink", revlink)
                mailstatus = SendEmailAct(fromaddr,email,login,password,host,port,cantext,'Candidate Email')
                print(mailstatus)
        elif(type=="nominee"):
            nomstatus = results[9][1]
            if(nomstatus == 'true'):
                nomtext = results[10][1]
                nomtext = str(nomtext)
                nomtext = nomtext.replace('@Name', name)
                nomtext = nomtext.replace('@Email', email)
                nomtext = nomtext.replace('@FormName', formname)
                nomtext = nomtext.replace('@SubmittedAt', submittedAt)
                nomtext = nomtext.replace('@InterviewLink', answerlink)
                nomtext = nomtext.replace('@FormLink', inteviewlink)
                nomtext = nomtext.replace("@ReviewFormLink", revlink)
                mailstatus = SendEmailAct(fromaddr,noemail,login,password,host,port,nomtext,'Nominee Email')
                print(mailstatus)
        elif(type=="reviewer"):
            revstatus = results[7][1]
            if(revstatus == 'true'):
                revtext = results[8][1]
                revtext = str(revtext)
                revtext = revtext.replace('@Name', name)
                revtext = revtext.replace('@Email', email)
                revtext = revtext.replace('@FormName', formname)
                revtext = revtext.replace('@SubmittedAt', submittedAt)
                revtext = revtext.replace('@InterviewLink', answerlink)
                revtext = revtext.replace('@FormLink', inteviewlink)
                revtext = revtext.replace("@ReviewFormLink", revlink)
                mailstatus = SendEmailAct(fromaddr,reveiweremail,login,password,host,port,revtext,'Reviewer Email')
                print(mailstatus)
    return "sent"     

def SendEmailAct(fromaddr,sendto,login,password,host, port,message,subject):
    print(sendto)
    acport = ''
    if(int(port) == 25):
        acport = int(587)
    else:
       acport = int(port)
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = fromaddr
    msg['To'] = sendto
    part2 = MIMEText(message, 'html')
    msg.attach(part2)
    s = smtplib.SMTP(host,int(acport))
    s.login(login, password)
    s.sendmail(fromaddr, sendto, msg.as_string())
    s.quit()
    return "mail Sent" 



@csrf_exempt
@require_http_methods(["POST"])
#@require_POST
def webhook_submit(request):
    anslist = []
    columnlist = []
    canName = ''
    canEmail = ''
    nomineemail = ''
    reviewemail = ''
    canNameval = ''
    canEmailval = ''
    nomineemailval = ''
    reviewemailval = ''
    anslinks = ''
    if(request.method=='GET'):
        str1 = "Get"
    elif(request.method=='POST'):
        studentid = ''
        json_body = json.loads(request.body)
        formId = json_body['formId']
        formName  = json_body['formName']
        formType  = json_body['interType']
        subAt = json_body['submittedAt']
        ip = json_body['ip']
        fields= json_body['fields']
        atblName = "ans_"+str(formId[-8:]).lower()
        connection =  mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        getcolumns= "SELECT `name`, `email`, `review`,`nominee` FROM form_column_mapping WHERE tblName = '"+atblName.strip()+"';"
        cursor = connection.cursor()
        cursor.execute(getcolumns)
        results = cursor.fetchall()
        for i, t in enumerate(results):
            if(i ==0):
                canName =t[0] 
                canEmail =t[1]
                nomineemail =t[2] 
                reviewemail =t[3]
        for i, f in enumerate(fields):
            ques_type   =  f['type']
            field   =  f['field']
            name = f['question']
            value   =  f['answer']
            if(ques_type == "link"):
                if(len(value) !=0):
                    anslinks += "<br> "+ name + " <br> View Answer : <a href="+ value + " target='_blank'> "+ value + " </a>"
                else:
                    anslinks += "<br> "+ name + " <br> View Answer : <span><strong>Not Applicable</strong></span>"
                anslist.append(value)
                columnlist.append(field)
            else:
                anslist.append(value)
                columnlist.append(field)
            if(canName == field):
                canNameval = value
            elif(canEmail == field):
                canEmailval = value
            elif(nomineemail == field):
                nomineemailval = value
            elif(reviewemail == field):
                reviewemailval = value
        anslist.append(ip)
        columnlist.append('ipAddress')
        anslist.append(subAt)
        columnlist.append('SubDate')
        anslist.append('No')
        columnlist.append('reviewStatus')
        anslist.append('No')
        columnlist.append('sent')
        anslist.append('No')
        columnlist.append('sentReport')
        anslist.append('No')
        columnlist.append('overGrade')
        instrow = generateInsertSchema(atblName,columnlist,anslist)
        cursor.execute(instrow)
        connection.commit()
        studentid = cursor.lastrowid
        getuserid= "SELECT * FROM forms_templates WHERE tbl_Name = '"+atblName.strip()+"';"
        cursor = connection.cursor()
        cursor.execute(getuserid)
        results = cursor.fetchone()
        if(len(canEmailval)!= 0):
            if(results[25] == 'no'):
                sendEmailinterview(results[8],'candidate',canNameval,canEmailval,formName,formId,subAt,'',atblName,anslinks,nomineemailval,reviewemailval)
            else:
                arraylist = []
                for i, f in enumerate(fields):
                    if(f['question'] != "AutoGrade"):
                        namelist = f['question']
                        valuelist   =  f['answer']
                        data = {}
                        data['ques'] = re.sub(r'[?|$|.|!]',r'', str(namelist).strip())
                        data['ans'] = re.sub(r'[?|$|.|!]',r'', str(valuelist).strip())
                        arraylist.append(data)
                formLink = 'https://test.techdivaa.com/'+formId
                reviewLink = 'https://test.techdivaa.com/review/'+str(studentid)+'/'+atblName+''
                reportstatus = sendpdfemail(arraylist,str(studentid),atblName,results[18],formId,formType,canNameval,canEmailval,formName,subAt,formLink,reviewLink,str(results[8]))
                if(reportstatus == "mail Sent"):
                    cursor10 = connection.cursor()
                    srptuptmt = 'UPDATE '+atblName+' SET sentReport = "yes" WHERE '+atblName+'id ='+str(studentid)+';'
                    cursor10.execute(srptuptmt)
                    connection.commit()
                    cursor10.close()
        if(len(nomineemailval)!= 0):
            sendEmailinterview(results[8],'nominee',canNameval,canEmailval,formName,formId,subAt,'',atblName,anslinks,nomineemailval,reviewemailval)
        if(len(reviewemailval)!= 0):
            sendEmailinterview(results[8],'reviewer',canNameval,canEmailval,formName,formId,subAt,'',atblName,anslinks,nomineemailval,reviewemailval)
        cursor.close()
    return HttpResponse(status=200)

class profileView(View):
    return_url = None
    def get(self , request,Id=''):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
          userdetails = RegistrationTbl.objects.get(email=userEmail.strip())
          planDetails = PlanTable.objects.get(name =userdetails.Plan)
          planlist = str(planDetails.features).split(',');
          return render(request , 'profile.html',{'data': userdetails,'PlanDetails':planlist}) 
        else:
            return redirect('login')

class teamView(View):
    return_url = None
    def get(self , request,Id=''):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
           return render(request , 'team.html') 
        else:
            return redirect('login')

class aifeedbackView(View):
    return_url = None
    def get(self , request,Id=''):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
           return render(request , 'AIfeedback.html') 
        else:
            return redirect('login')
    # Create your views here.
class formView(View):
    return_url = None
    def get(self , request,Id=''):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
            fid = Id.split("~")
            connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
            cursor = connection.cursor()
            cursor.execute("SELECT t.*, m.name as name, m.email as email, m.review as review, m.nominee as nominee FROM `forms_templates` as t INNER join `form_column_mapping` as m on t.tbl_Name= m.tblName WHERE t.form_ID='"+str(fid[0])+"'")
            rows = cursor.fetchone()
            cursor.close()
            return render(request , 'form.html',{'id':rows[0],'formName':rows[1],'formID':rows[2],'status':rows[11],'url':rows[9],'timer':rows[12],'imgurl':rows[6],'public':rows[14],'foldId':rows[13],'formlocode':rows[17],'formloposition':rows[21],'formlostatus':rows[20],'negativemarking':rows[22],'f_question_type':rows[19],'f_auto_report':rows[25],'videoc_collob':rows[15],'map_name':rows[26],'map_email':rows[27],'map_reviewer':rows[28],'map_nominee':rows[29]})
        else:
            return redirect('login')

    # Create your views here.
class workspaceView(View):
    return_url = None
    def get(self , request,Id=0):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
            return render(request , 'workspace.html')
        else:
            return redirect('login')

class studentView(View):
    return_url = None
    def get(self , request,Id=0):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
            return render(request , 'students.html')
        else:
            return redirect('login')

class studentListView(View):
    return_url = None
    def get(self , request,Id=0):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
            return render(request , 'studentslist.html')
        else:
            return redirect('login')

class interviewTempView(View):
    return_url = None
    def get(self , request,Id=0):
        userId = request.session.get('userId')
        userEmail = request.session.get('userEmail')
        if(userId!=None):
            return render(request , 'Interview_TemplatePreview.html', {'templateName': str(Id)+ '.html'})
        else:
            return redirect('login')
            
@csrf_exempt
def form_creationview(request):
    if request.method == 'POST':
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        utmpName = request.POST.get("tmpName")
        utmpType = request.POST.get("tmpType")
        utmpImg = request.POST.get("tmpImg")
        utmpId = request.POST.get("tmpId")
        ufolId = request.POST.get("folID")
        umarking = request.POST.get("nmarking")
        upublicView = request.POST.get("tmppublic")
        uintev_type = request.POST.get("inter_type")
        uautoreport = request.POST.get("autoreporting")
        id = generateId()
        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Forms/')
        fileName = "new_"+utmpName.replace(' ','_')
        fileName = fileName+'.png'
        file_path=fs.save(fileName,decode_base64_file(utmpImg)) 
        url = '/media/Forms/'+file_path
        userId = request.session.get('dbId')
        atblName = "ans_"+str(id[-8:]).lower()
        if(utmpType=='blank'):
            formtype = 'form'
            cursor = connection.cursor() 
            cursor.execute("INSERT INTO `forms_templates`( `form_Name`, `form_ID`,`tbl_Name`, `Parent_Template`, `form_img`, `formtype`,`user_Id`,`folder_id`,`public`,`negative_marking`,`form_Question_type`,`import_mode`,`auto_report`) VALUES ('"+utmpName+"','"+id+"','"+atblName+"','"+utmpType+"','"+url+"','"+formtype+"','"+str(userId)+"','"+ufolId+"','"+str(upublicView)+"','"+str(umarking)+"','"+str(uintev_type)+"','manual','"+str(uautoreport)+"')")
            connection.commit()
            cursor.close()
        elif(utmpType=='inteview'):
            cursor1 = connection.cursor()
            cursor1.execute("SELECT `id`, `tfieldName`, `tfieldData`, `tfieldsetData` FROM `templatefields` where `tformid`='"+utmpId+"';")
            results = list(cursor1.fetchall())
            cursor1.close()
            formtype = 'form'
            cursor2 = connection.cursor() 
            cursor2.execute("INSERT INTO `forms_templates`( `form_Name`, `form_ID`,`tbl_Name`, `Parent_Template`, `form_img`, `formtype`,`user_Id`,`folder_id`,`public`,`negative_marking`,`form_Question_type`,`import_mode`,`auto_report`) VALUES ('"+utmpName+"','"+id+"','"+atblName+"','"+utmpType+"','"+url+"','"+formtype+"','"+str(userId)+"','"+ufolId+"','"+str(upublicView)+"','"+str(umarking)+"','"+str(uintev_type)+"','manual','"+str(uautoreport)+"')")
            connection.commit()
            formid = cursor2.lastrowid
            cursor2.close()
            for r in results:
                fid = r[1]
                fidhtml = r[2]
                fidfieldset = r[3]
                cursor5 = connection.cursor()
                cursor5.execute("INSERT INTO `formfields`(`fieldName`, `fieldData`, `fieldsetData`, `formid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','"+str(formid)+"')")
                connection.commit()
                cursor5.close()
        elif(utmpType=='TalentSumo'):
            cursor1 = connection.cursor()
            cursor1.execute("SELECT `id`, `tfieldName`, `tfieldData`, `tfieldsetData` FROM `templatefields` where `tformid`='"+utmpId+"';")
            results = list(cursor1.fetchall())
            cursor1.close()
            formtype = 'form'
            cursor2 = connection.cursor() 
            cursor2.execute("INSERT INTO `forms_templates`( `form_Name`, `form_ID`,`tbl_Name`, `Parent_Template`, `form_img`, `formtype`,`user_Id`,`folder_id`,`public`,`negative_marking`,`form_Question_type`,`import_mode`,`auto_report`) VALUES ('"+utmpName+"','"+id+"','"+atblName+"','"+utmpType+"','"+url+"','"+formtype+"','"+str(userId)+"','"+ufolId+"','"+str(upublicView)+"','"+str(umarking)+"','"+str(uintev_type)+"','manual','"+str(uautoreport)+"')")
            connection.commit()
            formid = cursor2.lastrowid
            cursor2.close()
            for r in results:
                fid = r[1]
                fidhtml = r[2]
                fidfieldset = r[3]
                cursor5 = connection.cursor()
                cursor5.execute("INSERT INTO `formfields`(`fieldName`, `fieldData`, `fieldsetData`, `formid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','"+str(formid)+"')")
                connection.commit()
                cursor5.close()
        cursor7 = connection.cursor()
        cursor7.execute("CREATE TABLE "+atblName+" ("+atblName+"id INT(6) UNSIGNED AUTO_INCREMENT PRIMARY KEY, SubDate Text NULL,ipAddress Text NULL,commReview Text NULL,videoColloboration Text NULL,reviewStatus Text NULL,overGrade Text NULL,sent Text NULL,sentDate Text NULL,sentReport Text NULL,ReportLink Text NULL,AutoGrade Text NULL );")
        connection.commit()
        cursor7.execute("INSERT INTO "+atblName+"(`SubDate`,`ipAddress`,`commReview`,`videoColloboration`,`reviewStatus`,`overGrade`,`sent`,`sentDate`,`ReportLink`,`AutoGrade`) VALUES ('','','','','','','','','','');")
        connection.commit()
        cursor7.execute('INSERT INTO `form_column_mapping`( `formid`, `tblName`, `name`, `email`, `review`, `nominee`) VALUES ("'+id+'","'+atblName+'","","","","");')
        connection.commit()
        cursor7.close()
        return HttpResponse(id)

def generateId():
    id= ''
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    for j in range(0,10):
        res = ''.join(secrets.choice(string.ascii_uppercase +string.digits+ string.ascii_lowercase + string.digits + string.ascii_uppercase) for i in range(32))
        cursor.execute("SELECT COUNT(*) FROM `forms_templates` WHERE form_ID="+'"'+res+'"')
        rowcount = cursor.fetchone()
        j = j+1
        if rowcount[0]==0:  
            id = res
            break   
    cursor.close()  
    return  id
@csrf_exempt
def Imgupload_view(request):
    if request.method == 'POST':
        url = ''
        files = request.FILES.getlist("file") 
        #storage_client = storage.Client.from_service_account_json('/home/techtvxs/Interview/InterviewApp/cloud_json.json')
        #bucket_name = "interview_app_images"
        #bucket = storage_client.get_bucket(bucket_name)
        if len(files) != 0:
            for file in files:
                #blob = bucket.blob(str(file.name.replace(' ','_')))
                #blob.upload_from_file(file)
                fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/images/')
                file_path=fs.save(file.name.replace(' ','_'),file) 
                #blob.make_public()
                #url = blob.public_url
                url ='/media/images/'+ file_path
            else:
                print('No File') 
        return HttpResponse(url)

@csrf_exempt
def PDFlogo_Imgupload_view(request):
    if request.method == 'POST':
        url = ''
        files = request.FILES.getlist("file") 
        ufileName = request.POST.get("fileName")
        #storage_client = storage.Client.from_service_account_json('/home/techtvxs/Interview/InterviewApp/cloud_json.json')
        #bucket_name = "interview_app_images"
        #bucket = storage_client.get_bucket(bucket_name)
        if len(files) != 0:
            for file in files:
                #blob = bucket.blob(str(file.name.replace(' ','_')))
                #blob.upload_from_file(file)
                fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/report_logos/Clients/')
                file_path=fs.save(ufileName,file) 
                #blob.make_public()
                #url = blob.public_url
                url ='media/report_logos/Clients/'+ file_path
            else:
                print('No File') 
        return HttpResponse(url)

        
    
@csrf_exempt
def AnsTemplupload_view(request):
    if request.method == 'POST':
        url = ''
        files = request.FILES.getlist("file") 
        uformId = request.POST.get("formid")
        uformName = request.POST.get("formName")
        ufileName = request.POST.get("fileName")
        #storage_client = storage.Client.from_service_account_json('/home/techtvxs/Interview/InterviewApp/cloud_json.json')
        #bucket_name = "interview_app_images"
        #bucket = storage_client.get_bucket(bucket_name)
        if len(files) != 0:
            for file in files:
                #blob = bucket.blob(str(file.name.replace(' ','_')))
                #blob.upload_from_file(file)
                fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Answer_Templates/')
                file_path=fs.save(ufileName.replace(' ','_'),file) 
                #blob.make_public()
                #url = blob.public_url
                url ='/media/Answer_Templates/'+ file_path
                connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
                cursor = connection.cursor()
                cursor.execute("UPDATE `forms_templates` SET `answer_link`='"+str(url)+"' WHERE `form_ID` ='"+str(uformId)+"'")
                connection.commit()
                cursor.close()
                url = 'sucess'
            
        else:
            print('No File') 
            url = 'No File'
        return HttpResponse(url)



@csrf_exempt
def publish_pageview(request):
    if request.method == 'POST':
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        ucontents = request.POST.get("publishcontent")
        uformName = request.POST.get("formName")
        uvideo = request.POST.get("video")
        uscore_sttngs = request.POST.get("score_setting")
        uformId = request.POST.get("formID")
        uftimer = request.POST.get("timer")
        ufield_columns = request.POST['field_columns']
        ufield_columns_data = json.loads(ufield_columns)
        ufirst_columns = request.POST['first_columns']
        ufirst_columns_data = json.loads(ufirst_columns)
        umap_name = request.POST.get("map_name")
        umap_email = request.POST.get("map_email")
        umap_nominee = request.POST.get("map_nominee")
        umap_reviewer = request.POST.get("map_reviewer")
        filelocation = '/home/techtvxs/Forms/formApps/templates/'
        #filelocation = r'D:\\Python\\Interview New Backup\\Interview\\Interview\\InterviewApp\\templates\\'
        with open((filelocation +uformId+'.html'), 'w+', encoding="utf-8") as static_file:
            static_file.write(ucontents)
        cursor = connection.cursor()
        url = 'https://spellcheck.techdivaa.com/'+uformId
        #url = 'http://127.0.0.1:8000/pform/'+uformId
        date = datetime.datetime.now()
        cursor.execute("UPDATE `forms_templates` SET `published_url`='"+url+"',`timer`='"+uftimer+"',`published_Date`='"+str(date)+"',`published_Status`=1,`videocolloboration`='"+str(uvideo)+"',`score_settings`='"+str(uscore_sttngs)+"' WHERE form_ID='"+uformId+"'")
        connection.commit()
        atblName = "ans_"+str(uformId[-8:]).lower()
        ufield_columns_data.insert(0, atblName + "id")
        ufirst_columns_data.insert(0, "1")
        ufield_columns_data.insert(len(ufield_columns_data), "SubDate")
        ufirst_columns_data.insert(len(ufield_columns_data), "SubDate")
        ufield_columns_data.insert(len(ufield_columns_data)+1, "ipAddress")
        ufirst_columns_data.insert(len(ufield_columns_data)+1, "ipAddress")
        ufield_columns_data.insert(len(ufield_columns_data)+2,  "commReview")
        ufirst_columns_data.insert(len(ufield_columns_data)+2, "commReview")
        ufield_columns_data.insert(len(ufield_columns_data)+3, "videoColloboration")
        ufirst_columns_data.insert(len(ufield_columns_data)+3, "videoColloboration")
        ufield_columns_data.insert(len(ufield_columns_data)+4,  "reviewStatus")
        ufirst_columns_data.insert(len(ufield_columns_data)+4,  "reviewStatus")
        ufield_columns_data.insert(len(ufield_columns_data) +5, "overGrade")
        ufirst_columns_data.insert(len(ufield_columns_data) +5, "overGrade")
        ufield_columns_data.insert(len(ufield_columns_data) +6,"sent")
        ufirst_columns_data.insert(len(ufield_columns_data) +6,"sent")
        ufield_columns_data.insert(len(ufield_columns_data)+7,  "sentDate")
        ufirst_columns_data.insert(len(ufield_columns_data)+7,  "sentDate")
        ufield_columns_data.insert(len(ufield_columns_data) +8,"sentReport")
        ufirst_columns_data.insert(len(ufield_columns_data) +8,"sentReport")
        ufield_columns_data.insert(len(ufield_columns_data) +9, "ReportLink")
        ufirst_columns_data.insert(len(ufield_columns_data) +9, "ReportLink")
        ufield_columns_data.insert(len(ufield_columns_data) +8, "AutoGrade")
        ufirst_columns_data.insert(len(ufield_columns_data) +8, "AutoGrade")
        updateColumMapping = "UPDATE `form_column_mapping` SET `name`='"+umap_name+"',`email`='"+umap_email+"',`review`='"+umap_reviewer+"',`nominee`='"+umap_nominee+"' WHERE `formid`='"+str(uformId)+"' and tblName='"+str(atblName)+"';"
        cursor.execute(updateColumMapping)
        connection.commit()
        getcolumnli  = getcolumns(atblName)
        difqus = list(set(ufield_columns_data) - set(getcolumnli))
        if(len(difqus) > 0):
            for d in difqus:
                dropstement = 'ALTER TABLE '+atblName+' ADD '+str(d)+' VARCHAR(800);'
                cursor1 = connection.cursor()
                cursor1.execute(dropstement)
                connection.commit()
                cursor1.close()
        elif(len(difqus) == 0):
             difqus1 = list(set(getcolumnli) - set(ufield_columns_data))
             if(len(difqus1) > 0):
                for d in difqus1:
                    dropstement = 'ALTER TABLE '+atblName+' DROP COLUMN '+str(d)+' VARCHAR(800);'
                    cursor1 = connection.cursor()
                    cursor1.execute(dropstement)
                    connection.commit()
                    cursor1.close()
             else:
                pass
        updatecolumns =  'UPDATE '+atblName+' SET '
        for i, (field,first) in enumerate(zip(ufield_columns_data,ufirst_columns_data)):
            updatecolumns+= str(field)+" = '"+str(first) + "',"
        updatecolumns=updatecolumns[:-1]
        updatecolumns+=' WHERE '+atblName+'id =1;'
        cursor2 = connection.cursor()
        cursor2.execute(updatecolumns)
        connection.commit()
        cursor2.close()
        cursor.close()
        return HttpResponse('https://spellcheck.techdivaa.com/'+uformId)
        #return HttpResponse('http://127.0.0.1:8000/pform/'+uformId)




class page_view(View):
    return_url = None
    def get(self , request,Id=''):
        if(len(Id) != 0):
            connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
            cursor = connection.cursor()
            cursor.execute("SELECT published_Status FROM `forms_templates` WHERE form_ID='"+Id+"'")
            rowcount = cursor.fetchone()
            cursor.close()
            if(rowcount[0]==0):
                return render(request, "page.html", {'templateName': 'inactivetemplate.html'})
            else:
                return render(request, "page.html", {'templateName': str(Id)+ '.html'})
        else:
            return render(request, "page.html",{'templateName': 'blank.html'})

@csrf_exempt
def unpublish_pageview(request):
    if request.method == 'POST':
        uformId = request.POST.get("formID")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("UPDATE `forms_templates` SET `published_Status`=0 WHERE form_ID='"+uformId+"'")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')
        
@csrf_exempt
@require_http_methods(["POST"])
#@require_POST
def InsertVideo(request):
    if request.method =="POST":
        file = request.FILES['file']
        fname = request.POST.get('filename')
        ftype = request.POST.get('type')
        url=''
        storage_client = storage.Client.from_service_account_json('/home/techtvxs/Interview/InterviewApp/interviewapp.json')
        if ftype=='audio':
            bucket_name = "interview_app_audio"
            bucket = storage_client.get_bucket(bucket_name)
            #fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/audios/')
            #file_path=fs.save(fname.replace(' ','_'),file) 
            blob = bucket.blob(str(fname.replace(' ','_')))
            blob.upload_from_file(file)
            blob.make_public()
            url = blob.public_url
        else:
            #fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/videos/')
            #file_path=fs.save(fname.replace(' ','_'),file) 
            bucket_name = "interview_app_videos"
            bucket = storage_client.get_bucket(bucket_name)
            blob = bucket.blob(str(fname.replace(' ','_')))
            blob.upload_from_file(file)
            blob.make_public()
            url = blob.public_url
        return HttpResponse(url)
        

@csrf_exempt
def update_fields_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        ucount = request.POST.get("fcount")
        ucform = request.POST.get("currentId")
        image = request.POST.get("image")
        udata = request.POST['fields']
        data = json.loads(udata)
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor1 = connection.cursor()
        upchkflds = "DELETE FROM `formfields` WHERE `formid`= '"+uformid+"';"
        cursor1.execute(upchkflds)
        connection.commit()
        for i, d in enumerate(data):
            fid = d['fieldid']
            fidhtml = d['fieldhtml']
            fidfieldset = d['fieldset']
            cursor1.execute("INSERT INTO `formfields`(`fieldName`, `fieldData`, `fieldsetData`, `formid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','"+uformid+"')")
            connection.commit()   
        cursor1.close()
        chkImg = "SELECT form_Name,form_img FROM `forms_templates` WHERE id = '"+uformid+"';"
        cursor5 = connection.cursor()
        cursor5.execute(chkImg)
        imgresults = cursor5.fetchone()
        imglength = ''.join(imgresults[1])
        formName = ''.join(imgresults[0])
        cursor5.close()
        if "new_" in imglength:
            if((int(ucount) >1) and (ucform!="welcomestep")):
                fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Forms/')
                fileName = formName.replace(' ','_')
                fileName = fileName+'.png'
                file_path=fs.save(fileName,decode_base64_file(image)) 
                url = '/media/Forms/'+file_path
                cursor4 = connection.cursor()
                cursor4.execute("UPDATE `forms_templates` SET `form_img`='"+str(url)+"' WHERE id = '"+uformid+"';")
                connection.commit()
                cursor4.close()
        return HttpResponse(imglength)

@csrf_exempt
def single_fields_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        udata = request.POST['fields']
        data = json.loads(udata)
        fid = data['fieldid']
        fidhtml = data['fieldhtml']
        fidfieldset = data['fieldset']
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        upchk = "SELECT `fieldName` FROM `formfields` WHERE `fieldName`='"+str(fid)+"' and `formid`= '"+uformid+"';"
        uprows =  cursor.execute(upchk)
        uprowscount = cursor.fetchone()
        if(uprowscount!=None):
            updaterow = "UPDATE `formfields` SET `fieldData`='"+str(fidhtml)+"',`fieldsetData`='"+str(fidfieldset)+"' WHERE `fieldName`='"+str(fid)+"' and `formid`= '"+uformid+"';"
            cursor.execute(updaterow)
            connection.commit()
        else:
            cursor.execute("INSERT INTO `formfields`(`fieldName`, `fieldData`, `fieldsetData`, `formid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','"+uformid+"')")
            connection.commit()
        cursor.close()
        return HttpResponse('Updated')

@csrf_exempt
def trigger_fields_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        udata = request.POST['fields']
        data = json.loads(udata)
        fid = data['fieldid']
        fidhtml = data['fieldhtml']
        fidfieldset = data['fieldset']
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        upchk = "SELECT `fieldName` FROM `formfields` WHERE `fieldName`='"+str(fid)+"' and `formid`= '"+uformid+"';"
        uprows =  cursor.execute(upchk)
        uprowscount = cursor.fetchone()
        if(uprowscount!=None):
            updaterow = "UPDATE `formfields` SET `fieldData`='"+str(fidhtml)+"',`fieldsetData`='"+str(fidfieldset)+"' WHERE `fieldName`='"+str(fid)+"' and `formid`= '"+uformid+"';"
            cursor.execute(updaterow)
            connection.commit()
        cursor.close()
        return HttpResponse('Updated')

@csrf_exempt
def get_fields_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt = "SELECT `fieldName`,`fieldData`,`fieldsetData` FROM `formfields` WHERE `formid`='"+str(uformid)+"';"
        cursor.execute(stmt)
        datarows = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows,safe=False)
        
@csrf_exempt
def get_tempformfields_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt = "SELECT `tfieldName`,`tfieldData`,`tfieldsetData` FROM `templatefields` WHERE `tformid`='"+str(uformid)+"';"
        cursor.execute(stmt)
        datarows = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows,safe=False)

@csrf_exempt
def get_Temp_fields_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("tempId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt = "SELECT `tfieldName`,`tfieldData`,`tfieldsetData` FROM `templatefields` WHERE `tformid`='"+str(uformid)+"';"
        cursor.execute(stmt)
        datarows = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows,safe=False)

@csrf_exempt
def getForms_View(request):
    if request.method == 'POST':
        responses = []
        ufolid = request.POST.get("FolderId")
        userId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt = "SELECT * FROM `forms_templates` where folder_id='"+str(ufolid)+"' and user_Id ='"+str(userId)+"';"
        cursor.execute(stmt)
        datarows1 = list(cursor.fetchall())
        for tblName in datarows1:
            try:
                cursor.execute("SELECT count(*) FROM `"+tblName[3]+"`")
                rows1 = cursor.fetchone()
                responses.append(rows1[0])
            except:
                responses.append(0)
        cursor.close()
        responses = {"data":datarows1,"responses":responses}
        return JsonResponse(responses,safe=False)

@csrf_exempt
def get_sharedbyforms_View(request):
    if request.method == 'POST':
        userId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt = "SELECT * FROM `forms_templates` where public='1' and user_Id ='"+str(userId)+"';"
        cursor.execute(stmt)
        datarows1 = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows1,safe=False)

@csrf_exempt
def get_sharedwithforms_View(request):
    if request.method == 'POST':
        userId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("select member_of from tbl_member_mapping where userid= "+str(userId)+";")
        uidrows = cursor.fetchall()
        datarows1 = []
        if(uidrows !=[]):
            uids = []
            for i in uidrows:
                uidvalue = list(i)
                uids.append("'"+str(uidvalue[0])+"'")
            strnids=','.join([str(x) for x in uids])
            stmt = "SELECT * FROM `forms_templates` where public='1' and user_Id in ("+str(strnids)+");;"
            cursor.execute(stmt)
            datarows1 = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows1,safe=False)

@csrf_exempt
def getTemp_View(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt =  "SELECT * FROM `templ_tbl` where temp_type IN ('blank','inteview');"
        cursor.execute(stmt)
        datarows = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows,safe=False)

@csrf_exempt
def getTalentTemp_View(request):
    if request.method == 'POST':
        utfolder = request.POST.get("FolderId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt =  "SELECT * FROM `templ_tbl` where folder_id ='"+utfolder+"';"
        cursor.execute(stmt)
        datarows = list(cursor.fetchall())
        datarows1 = []
        if(len(datarows)!=0):
            for id in datarows:
                stmt1 =  "SELECT * FROM `templatefields` where tformid  ='"+str(id[3])+"';"
                cursor.execute(stmt1)
                rows = list(cursor.fetchall())
                if(len(rows)!=0):
                    stmt2 =  "SELECT * FROM `templ_tbl` where id ='"+str(id[3])+"';"
                    cursor.execute(stmt2)
                    rowdata = cursor.fetchone()
                    datarows1.append(rowdata)
        cursor.close()
        return JsonResponse(datarows1,safe=False)

@csrf_exempt
def publishTalentTemplate_view(request):
    if request.method == 'POST':
        image = request.POST.get("image")
        templId = request.POST.get("TempID")
        ltemplId = request.POST.get("LongTempID")
        tcontents = request.POST.get("publishcontent")
        udata = request.POST['fields']
        data = json.loads(udata)
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        getformdtls = "SELECT * FROM  templatefields WHERE tformid ='"+str(templId)+"';"
        cursor.execute(getformdtls)
        results = cursor.fetchone()
        allrows = cursor.fetchall()
        if(results== None):
            for i, d in enumerate(data):
                fid = d['fieldid']
                fidhtml = d['fieldhtml']
                fidfieldset = d['fieldset']
                cursor.execute("INSERT INTO `templatefields`(`tfieldName`, `tfieldData`, `tfieldsetData`, `tformid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','"+str(templId)+"')")
                connection.commit()
        else:
            cursor.execute("DELETE FROM `templatefields` where `tformid`='"+str(templId)+"'")
            connection.commit()
            for i, d in enumerate(data):
                fid = d['fieldid']
                fidhtml = d['fieldhtml']
                fidfieldset = d['fieldset']
                cursor.execute("INSERT INTO `templatefields`(`tfieldName`, `tfieldData`, `tfieldsetData`, `tformid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','"+str(templId)+"')")
                connection.commit()
        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Templates/')
        fileName = 'Sample Interview'.replace(' ','_')
        fileName = fileName+'.png'
        file_path=fs.save(fileName,decode_base64_file(image)) 
        url = '/media/Templates/'+file_path
        cursor.execute("UPDATE `templ_tbl` SET `temp_img`='"+str(url)+"' WHERE `id`= '"+str(templId)+"';")
        connection.commit()
        cursor.close()
        filelocation = '/home/techtvxs/Interview/InterviewApp/templates/'
        with open((filelocation +ltemplId+'.html'), 'w+', encoding="utf-8") as static_file:
            static_file.write(tcontents)
        return HttpResponse('success')



@csrf_exempt
def publishTemplate_view(request):
    if request.method == 'POST':
        image = request.POST.get("image")
        udata = request.POST['fields']
        data = json.loads(udata)
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        getformdtls = "SELECT * FROM  templatefields WHERE tformid ='3';"
        cursor.execute(getformdtls)
        results = cursor.fetchone()
        allrows = cursor.fetchall()
        cursor.close()
        if(results== None):
            for i, d in enumerate(data):
                fid = d['fieldid']
                fidhtml = d['fieldhtml']
                fidfieldset = d['fieldset']
                cursor1 = connection.cursor()
                cursor1.execute("INSERT INTO `templatefields`(`tfieldName`, `tfieldData`, `tfieldsetData`, `tformid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','3')")
                connection.commit()
                cursor1.close()
        else:
            cursor2 = connection.cursor()
            cursor2.execute("DELETE FROM `templatefields` where `tformid`=3")
            connection.commit()
            cursor2.close()
            for i, d in enumerate(data):
                fid = d['fieldid']
                fidhtml = d['fieldhtml']
                fidfieldset = d['fieldset']
                cursor3 = connection.cursor()
                cursor3.execute("INSERT INTO `templatefields`(`tfieldName`, `tfieldData`, `tfieldsetData`, `tformid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','3')")
                connection.commit()
                cursor3.close()
        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Templates/')
        fileName = 'Sample Interview'.replace(' ','_')
        fileName = fileName+'.png'
        file_path=fs.save(fileName,decode_base64_file(image)) 
        url = '/media/Templates/'+file_path
        cursor4 = connection.cursor()
        cursor4.execute("UPDATE `templ_tbl` SET `temp_img`='"+str(url)+"' WHERE `id`= 3;")
        connection.commit()
        cursor4.close()
        return HttpResponse('success')




def decode_base64_file(data):
    def get_file_extension(file_name, decoded_file):
        import imghdr
        extension = imghdr.what(file_name, decoded_file)
        extension = "jpg" if extension == "jpeg" else extension
        return extension
    # Check if this is a base64 string
    if isinstance(data, six.string_types):
        # Check if the base64 string is in the "data:" format
        if 'data:' in data and ';base64,' in data:
            # Break out the header from the base64 content
            header, data = data.split(';base64,')

        # Try to decode the file. Return validation error if it fails.
        try:
            decoded_file = base64.b64decode(data)
        except TypeError:
            TypeError('invalid_image')

        # Generate file name:
        file_name = str(uuid.uuid4())[:12] # 12 characters are more than enough.
        # Get the file name extension:
        file_extension = get_file_extension(file_name, decoded_file)

        complete_file_name = "%s.%s" % (file_name, file_extension, )

        return ContentFile(decoded_file, name=complete_file_name)

@csrf_exempt
def update_form_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        uName = request.POST.get("formName")
        uFoldId = request.POST.get("folID")
        upublic = request.POST.get("tmppublic")
        umarking = request.POST.get("nmarking")
        uautoreport = request.POST.get("autoreporting")
        uformType = request.POST.get("inter_type")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("UPDATE `forms_templates` SET `form_Name`='"+str(uName)+"',`folder_id`='"+str(uFoldId)+"',`public`='"+str(upublic)+"',`negative_marking`='"+str(umarking)+"',`form_Question_type`='"+str(uformType)+"',`auto_report`='"+str(uautoreport)+"' WHERE id ='"+str(uformid)+"'")
        connection.commit()
        cursor.close()
        return HttpResponse("success")

@csrf_exempt
def del_form_view(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        uformlongid = request.POST.get("formlongID")
        atblName = "ans_"+str(uformlongid[-8:]).lower()
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("DELETE FROM `formfields` WHERE formid='"+str(uformid)+"'")
        connection.commit()
        cursor.execute("DELETE FROM `forms_templates` WHERE id='"+str(uformid)+"'")
        connection.commit()
        cursor.execute("DELETE FROM `form_column_mapping` WHERE tblName='"+str(atblName)+"' and formid='"+str(uformlongid)+"'")
        connection.commit()
        cursor1 = connection.cursor()
        stmt = "SHOW TABLES LIKE '"+str(atblName)+"'"
        cursor1.execute(stmt)
        result = cursor1.fetchone()
        cursor1.close()
        if result:
            cursor.execute("DROP TABLE "+str(atblName)+";")
            connection.commit()
        cursor.close()
        return HttpResponse("success")
	           
@csrf_exempt
def duplicate_formview(request):
    if request.method == 'POST':
        uformid = request.POST.get("formId")
        uformlongid = request.POST.get("formLongID")
        uformName = request.POST.get("FormName")
        uformPublic = request.POST.get("FPublic")
        ufoldid = request.POST.get("FoldID")
        uformType = request.POST.get("inter_type")
        unegative = request.POST.get("nmarking")
        uautoreport = request.POST.get("autoreporting")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM `forms_templates` WHERE id="+'"'+uformid+'"')
        rows = cursor.fetchone()
        formName = rows[1]+ " - copy"
        id = generateId()
        atblName = "ans_"+str(id[-8:]).lower()
        formImg = rows[6]
        userId = request.session.get('dbId')
        cursor.execute("INSERT INTO `forms_templates`( `form_Name`, `form_ID`, `tbl_Name`, `Parent_Template`, `form_img`, `formtype`,`user_Id`,`folder_id`,`public`,`Duplicate`,`negative_marking`,`form_Question_type`,`auto_report`) VALUES ('"+uformName+"','"+id+"','"+atblName+"','blank','"+formImg+"','form','"+str(userId)+"','"+ufoldid+"','"+str(uformPublic)+"','1','"+str(unegative)+"','"+str(uformType)+"','"+str(uautoreport)+"')")
        connection.commit()
        formidv = cursor.lastrowid
        cursor.close()
        cursor1 = connection.cursor()
        cursor1.execute("SELECT `id`, `fieldName`, `fieldData`, `fieldsetData` FROM `formfields` where `formid`='"+uformid+"';")
        results = list(cursor1.fetchall())
        cursor1.close()
        for r in results:
            fid = r[1]
            fidhtml = r[2]
            fidfieldset = r[3]
            cursor5 = connection.cursor()
            cursor5.execute("INSERT INTO `formfields`(`fieldName`, `fieldData`, `fieldsetData`, `formid`) VALUES ('"+str(fid)+"','"+str(fidhtml)+"','"+str(fidfieldset)+"','"+str(formidv)+"')")
            connection.commit()
            cursor5.close()
        cursor7 = connection.cursor()
        cursor7.execute("CREATE TABLE "+atblName+" ("+atblName+"id INT(6) UNSIGNED AUTO_INCREMENT PRIMARY KEY, SubDate Text NULL,ipAddress Text NULL,commReview Text NULL,videoColloboration Text NULL,reviewStatus Text NULL,overGrade Text NULL,sent Text NULL,sentDate Text NULL,sentReport Text NULL,ReportLink Text NULL,AutoGrade Text NULL );")
        connection.commit()
        cursor7.execute("INSERT INTO "+atblName+"(`SubDate`,`ipAddress`,`commReview`,`videoColloboration`,`reviewStatus`,`overGrade`,`sent`,`sentDate`,`ReportLink`,`AutoGrade`) VALUES ('','','','','','','','','','');")
        connection.commit()
        cursor7.execute('INSERT INTO `form_column_mapping`( `formid`, `tblName`, `name`, `email`, `review`, `nominee`) VALUES ("'+id+'","'+atblName+'","","","","");')
        connection.commit()
        cursor7.close()
        return HttpResponse("success")

@csrf_exempt
def folder_creationview(request):
    if request.method == 'POST':
        ufolName = request.POST.get("folName")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor5 = connection.cursor()
        cursor5.execute("INSERT INTO `foldername`(`folderName`,`user_Id`) VALUES ('"+str(ufolName)+"','"+str(userdbId)+"')")
        connection.commit()
        cursor5.close()
        return HttpResponse("success")


@csrf_exempt
def get_foldersview(request):
    if request.method == 'POST':
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        userdbId = request.session.get('dbId')
        cursor = connection.cursor()
        stmt =  "SELECT * FROM `foldername` where `user_Id` ='"+str(userdbId)+"';"
        cursor.execute(stmt)
        datarows = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows,safe=False)

@csrf_exempt
def get_campaignlistview(request):
    if request.method == 'POST':
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        userdbId = request.session.get('dbId')
        cursor = connection.cursor()
        stmt =  "SELECT * FROM `tbl_students_lists` where `user_Id` ='"+str(userdbId)+"';"
        cursor.execute(stmt)
        datarows = list(cursor.fetchall())
        cursor.close()
        return JsonResponse(datarows,safe=False)

@csrf_exempt
def get_talentfoldersview(request):
    if request.method == 'POST':
        folders = TalenstsumoFolder.objects.all().values_list('id','name')
        data = []
        for f in folders:
            data.append({"id":f[0],"name":f[1]})
        return JsonResponse(json.dumps(data),safe=False)

@csrf_exempt
def update_foldersview(request):
    if request.method == 'POST':
        uformName = request.POST.get("folName")
        uformid = request.POST.get("folId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt =  "UPDATE `foldername` SET `folderName`='"+str(uformName)+"' WHERE id= '"+str(uformid)+"';"
        cursor.execute(stmt)
        connection.commit()
        cursor.close()
        return HttpResponse("success")


@csrf_exempt
def delete_foldersview(request):
    if request.method == 'POST':
        uformid = request.POST.get("folId")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt1 =  "DELETE FROM `forms_templates` WHERE folder_id ='"+str(uformid)+"';"
        stmt =  "DELETE FROM `foldername` WHERE id='"+str(uformid)+"';"
        cursor.execute(stmt1)
        connection.commit()
        cursor.execute(stmt)
        connection.commit()
        cursor.close()
        return HttpResponse("success")


class registerView(View):
    return_url = None
    def get(self , request,Id=0):
        return render(request , 'registration.html',{"email":'',"id":''})

class forgotView(View):
    return_url = None
    def get(self , request,Id=0):
            return render(request , 'forgot.html')
    @csrf_exempt
    def post(self , request,Id=0):
            return render(request, 'login.html')

class resendActivationView(View):
    return_url = None
    def get(self , request,Id=0):
            return render(request , 'resend_email.html')
    @csrf_exempt
    def post(self , request,Id=0):
            return render(request, 'login.html')

class resentPassword(View):
    return_url = None
    def get(self , request,Id=0):
            return render(request , 'reset_password.html')
    @csrf_exempt
    def post(self , request,Id=0):
            return render(request, 'login.html')


def checkemail(request):
   if request.method == 'GET':  
        uemail = request.GET['uEmail']
        ouemail =  RegistrationTbl.objects.filter(email=uemail.strip())
        if len(ouemail)!=0:
            result= "Exits"
        else:
            result= "Valid"
        return HttpResponse(result) 

@csrf_exempt
def pregister_view(request):
    if request.method == 'POST':
        ufullName = request.POST.get("uFullName")
        uemail = request.POST.get("uEmail")
        uid = request.POST.get("uId")
        upassword = request.POST.get("uPassword")
        if(len(uid)==0):
            register = RegistrationTbl(fullname=ufullName,email=uemail,password=upassword,is_active=0)
            register.save()
        else:
            register = RegistrationTbl.objects.get(email=uemail)
            register.fullname = ufullName
            register.password = upassword
            register.save()
        current_site = get_current_site(request)
        
        url = "http://" + current_site.domain + "/activate/" + str(urlsafe_base64_encode(force_bytes(register.pk))) + '/' + str(account_activation_token.make_token(register))
        sendActivationLink(uemail,url)
        return HttpResponse('sucess')

@csrf_exempt
def presend_view(request):
    if request.method == 'POST':
        uemail = request.POST.get("userEmail")
        register = RegistrationTbl.objects.get(email = uemail.strip())
        current_site = get_current_site(request)
        url = "http://" + current_site.domain + "/activate/" + str(urlsafe_base64_encode(force_bytes(register.pk))) + '/' + str(account_activation_token.make_token(register))
        sendActivationLink(uemail,url)
        return HttpResponse('sucess')


@csrf_exempt
def pforgot_view(request):
    if request.method == 'POST':
        uemail = request.POST.get("userEmail")
        register = RegistrationTbl.objects.get(email = uemail.strip())
        current_site = get_current_site(request)
        url = "http://" + current_site.domain + "/reset/" + str(urlsafe_base64_encode(force_bytes(register.pk))) + '/' + str(account_activation_token.make_token(register))
        sendResetLink(uemail,url)
        return HttpResponse('sucess')


@csrf_exempt
def plogin_View(request):
    if request.method == 'POST':
        uemail = request.POST.get("userEmail")
        upassword = request.POST.get("userPass")
        error_message =''
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        try:
            user = RegistrationTbl.objects.get(email=uemail.strip())
            if(user.is_active==1):
                cursor1 = connection.cursor()
                cursor1.execute("SELECT id FROM `registration_tbl` WHERE email="+'"'+uemail.strip()+'"')
                rows = cursor1.fetchone()
                cursor1.execute("SELECT count(`forms_templates`.`form_Name`) FROM `forms_templates` WHERE user_Id="+'"'+str(rows[0]).strip()+'"')
                rows1 = cursor1.fetchone()
                cursor1.close()
                if upassword == user.password:
                    request.session['userId'] = user.id
                    request.session['userEmail'] = user.email
                    request.session['userName'] = user.fullname
                    request.session['dbId'] = rows[0]
                    request.session['Plan'] = user.Plan.name
                    request.session['Number'] = rows1[0]
                    error_message = 'valid'
                else:
                    error_message = 'not valid'
            else:
                error_message = 'In-active'    
        except RegistrationTbl.DoesNotExist:
            error_message = 'not valid'
        return HttpResponse(error_message)

@csrf_exempt
def preset_View(request):
    if request.method == 'POST':
        uemail = request.POST.get("uEmail")
        upassword = request.POST.get("uPassword")
        register = RegistrationTbl.objects.get(email = uemail.strip())
        register.password = upassword
        register.save()
        return HttpResponse('sucess')

def sendActivationLink(uemail,url):
    mail_content = """
<!doctype html>
<html lang="en-US">
<head>
    <meta content="text/html; charset=utf-8" http-equiv="Content-Type" />
    <title>TalentSumo - Activate Account</title>
    <meta name="description" content="Activate Account TalentSumo.">
    <style type="text/css">
        a:hover {text-decoration: underline !important;}
    </style>
</head>

<body marginheight="0" topmargin="0" marginwidth="0" style="margin: 0px; background-color: #f2f3f8;" leftmargin="0">
    <table cellspacing="0" border="0" cellpadding="0" width="100%" bgcolor="#f2f3f8"
        style="@import url(https://fonts.googleapis.com/css?family=Rubik:300,400,500,700|Open+Sans:300,400,600,700); font-family: 'Open Sans', sans-serif;">
        <tr>
            <td>
                <table style="background-color: #f2f3f8; max-width:670px;  margin:0 auto;" width="100%" border="0"
                    align="center" cellpadding="0" cellspacing="0">
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                          <a href="https://www.talentsumo.co" title="logo" target="_blank">
                            <img width="250" src="https://i.ibb.co/Ssq0KWY/logo.png" title="logo" alt="logo">
                          </a>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td>
                            <table width="95%" border="0" align="center" cellpadding="0" cellspacing="0"
                                style="max-width:670px;background:#fff; border-radius:3px; text-align:center;-webkit-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);-moz-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);box-shadow:0 6px 18px 0 rgba(0,0,0,.06);">
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                                <tr>
                                    <td style="padding:0 35px;">
                                        <h1 style="color:#1e1e2d; font-weight:500; margin:0;font-size:32px;font-family:'Rubik',sans-serif;">TalentSumo Account Activation</h1>
                                        <span
                                            style="display:inline-block; vertical-align:middle; margin:29px 0 26px; border-bottom:1px solid #cecece; width:100%;"></span>
                                       
                                        <a href= """ + url + """
                                            style="background:black;text-decoration:none !important; font-weight:500; margin-top:10px;margin-bottom:10px; color:#fff;text-transform:uppercase; font-size:14px;padding:10px 24px;display:inline-block;border-radius:50px;">Activate
                                            Account</a>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="height:40px;color: black;">Note: above link is only valid for 24 Hours.</td>
                                </tr>
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                            </table>
                        </td>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                            <p style="font-size:14px; color:rgba(69, 80, 86, 0.7411764705882353); line-height:18px; margin:0 0 0;">&copy; <strong>www.talentsumo.co</strong></p>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>

</html>"""
    CommonMail(uemail,mail_content,'Activate your TalentSumo Account.')
    return "sent Mail"


def sendRegistrationLink(username,uemail,url):
    mail_content = """
<!doctype html>
<html lang="en-US">
<head>
    <meta content="text/html; charset=utf-8" http-equiv="Content-Type" />
    <title>TalentSumo - Team Invitation</title>
    <meta name="description" content="Team Invitation - TalentSumo.">
    <style type="text/css">
        a:hover {text-decoration: underline !important;}
    </style>
</head>

<body marginheight="0" topmargin="0" marginwidth="0" style="margin: 0px; background-color: #f2f3f8;" leftmargin="0">
    <table cellspacing="0" border="0" cellpadding="0" width="100%" bgcolor="#f2f3f8"
        style="@import url(https://fonts.googleapis.com/css?family=Rubik:300,400,500,700|Open+Sans:300,400,600,700); font-family: 'Open Sans', sans-serif;">
        <tr>
            <td>
                <table style="background-color: #f2f3f8; max-width:670px;  margin:0 auto;" width="100%" border="0"
                    align="center" cellpadding="0" cellspacing="0">
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                          <a href="https://www.talentsumo.co" title="logo" target="_blank">
                            <img width="250" src="https://i.ibb.co/Ssq0KWY/logo.png" title="logo" alt="logo">
                          </a>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td>
                            <table width="95%" border="0" align="center" cellpadding="0" cellspacing="0"
                                style="max-width:670px;background:#fff; border-radius:3px; text-align:center;-webkit-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);-moz-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);box-shadow:0 6px 18px 0 rgba(0,0,0,.06);">
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                                <tr>
                                    <td style="padding:0 35px;">
                                        <h1 style="color:#1e1e2d; font-weight:500; margin:0;font-size:32px;font-family:'Rubik',sans-serif;">TalentSumo - Team Invitation</h1>
                                        <span
                                            style="display:inline-block; vertical-align:middle; margin:29px 0 26px; border-bottom:1px solid #cecece; width:100%;"></span>
                                        <p style="color:#455056; font-size:15px;line-height:24px; margin:0;">
                                           """ + username.capitalize() + """ has invited to join there team.
                                        </p>
                                        <a href= """ + url + """
                                            style="background:black;text-decoration:none !important; font-weight:500; margin-top:10px;margin-bottom:10px; color:#fff;text-transform:uppercase; font-size:14px;padding:10px 24px;display:inline-block;border-radius:50px;">Join Now</a>
                                    </td>
                                </tr>
                                <tr>
                                    
                                </tr>
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                            </table>
                        </td>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                            <p style="font-size:14px; color:rgba(69, 80, 86, 0.7411764705882353); line-height:18px; margin:0 0 0;">&copy; <strong>www.talentsumo.co</strong></p>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>

</html>"""
    CommonMail(uemail,mail_content,'TalentSumo - Team Member Invitation.')
    return "sent Mail"
    
    
def sendMemberNotification(username,uemail):
    mail_content = """
<!doctype html>
<html lang="en-US">
<head>
    <meta content="text/html; charset=utf-8" http-equiv="Content-Type" />
    <title>TalentSumo - Team Member Details</title>
    <meta name="description" content="Team Invitation - TalentSumo.">
    <style type="text/css">
        a:hover {text-decoration: underline !important;}
    </style>
</head>

<body marginheight="0" topmargin="0" marginwidth="0" style="margin: 0px; background-color: #f2f3f8;" leftmargin="0">
    <table cellspacing="0" border="0" cellpadding="0" width="100%" bgcolor="#f2f3f8"
        style="@import url(https://fonts.googleapis.com/css?family=Rubik:300,400,500,700|Open+Sans:300,400,600,700); font-family: 'Open Sans', sans-serif;">
        <tr>
            <td>
                <table style="background-color: #f2f3f8; max-width:670px;  margin:0 auto;" width="100%" border="0"
                    align="center" cellpadding="0" cellspacing="0">
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                          <a href="https://www.talentsumo.co" title="logo" target="_blank">
                            <img width="250" src="https://i.ibb.co/Ssq0KWY/logo.png" title="logo" alt="logo">
                          </a>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td>
                            <table width="95%" border="0" align="center" cellpadding="0" cellspacing="0"
                                style="max-width:670px;background:#fff; border-radius:3px; text-align:center;-webkit-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);-moz-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);box-shadow:0 6px 18px 0 rgba(0,0,0,.06);">
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                                <tr>
                                    <td style="padding:0 35px;">
                                        <h1 style="color:#1e1e2d; font-weight:500; margin:0;font-size:32px;font-family:'Rubik',sans-serif;">TalentSumo - Team Details</h1>
                                        <span
                                            style="display:inline-block; vertical-align:middle; margin:29px 0 26px; border-bottom:1px solid #cecece; width:100%;"></span>
                                        <p style="color:#455056; font-size:15px;line-height:24px; margin:0;">
                                           """ + username.capitalize() + """ has added you within there team.
                                        </p>
                                        <a href= "https://test.techdivaa.com"
                                            style="background:black;text-decoration:none !important; font-weight:500; margin-top:10px;margin-bottom:10px; color:#fff;text-transform:uppercase; font-size:14px;padding:10px 24px;display:inline-block;border-radius:50px;">Login Now</a>
                                    </td>
                                </tr>
                                <tr>
                                    
                                </tr>
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                            </table>
                        </td>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                            <p style="font-size:14px; color:rgba(69, 80, 86, 0.7411764705882353); line-height:18px; margin:0 0 0;">&copy; <strong>www.talentsumo.co</strong></p>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>

</html>"""
    CommonMail(uemail,mail_content,'TalentSumo - Team Member Details.')
    return "sent Mail"


def sendResetLink(uemail,url):
    mail_content = """
<!doctype html>
<html lang="en-US">
<head>
    <meta content="text/html; charset=utf-8" http-equiv="Content-Type" />
    <title>TalentSumo - Reset Password</title>
    <meta name="description" content="Reset Password TalentSumo.">
    <style type="text/css">
        a:hover {text-decoration: underline !important;}
    </style>
</head>

<body marginheight="0" topmargin="0" marginwidth="0" style="margin: 0px; background-color: #f2f3f8;" leftmargin="0">
    <table cellspacing="0" border="0" cellpadding="0" width="100%" bgcolor="#f2f3f8"
        style="@import url(https://fonts.googleapis.com/css?family=Rubik:300,400,500,700|Open+Sans:300,400,600,700); font-family: 'Open Sans', sans-serif;">
        <tr>
            <td>
                <table style="background-color: #f2f3f8; max-width:670px;  margin:0 auto;" width="100%" border="0"
                    align="center" cellpadding="0" cellspacing="0">
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                          <a href="https://www.talentsumo.co" title="logo" target="_blank">
                            <img width="250" src="https://i.ibb.co/98RkDhx/techdivaa1.png" title="logo" alt="logo">
                          </a>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td>
                            <table width="95%" border="0" align="center" cellpadding="0" cellspacing="0"
                                style="max-width:670px;background:#fff; border-radius:3px; text-align:center;-webkit-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);-moz-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);box-shadow:0 6px 18px 0 rgba(0,0,0,.06);">
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                                <tr>
                                    <td style="padding:0 35px;">
                                        <h1 style="color:#1e1e2d; font-weight:500; margin:0;font-size:32px;font-family:'Rubik',sans-serif;">TalentSumo Reset Password</h1>
                                        <span
                                            style="display:inline-block; vertical-align:middle; margin:29px 0 26px; border-bottom:1px solid #cecece; width:100%;"></span>
                                        <p style="color:#455056; font-size:15px;line-height:24px; margin:0;">
                                            We cannot simply send you your old password. A unique link to reset your
                                            password has been generated for you. To reset your password, click the
                                            following link and follow the instructions.
                                        </p>
                                        <a href= """ + url + """
                                            style="background:black;text-decoration:none !important; font-weight:500; margin-top:10px;margin-bottom:10px; color:#fff;text-transform:uppercase; font-size:14px;padding:10px 24px;display:inline-block;border-radius:50px;">Reset
                                            Password</a>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="height:40px;color: black;">Note: above link is only valid for 24 Hours.</td>
                                </tr>
                                <tr>
                                    <td style="height:40px;"></td>
                                </tr>
                            </table>
                        </td>
                    <tr>
                        <td style="height:20px;"></td>
                    </tr>
                    <tr>
                        <td style="text-align:center;">
                            <p style="font-size:14px; color:rgba(69, 80, 86, 0.7411764705882353); line-height:18px; margin:0 0 0;">&copy; <strong>www.talentsumo.co</strong></p>
                        </td>
                    </tr>
                    <tr>
                        <td style="height:80px;"></td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>

</html>"""
    CommonMail(uemail,mail_content,'TalentSumo - Reset Password.')
    return "sent Mail"

def CommonMail (uemail, umessage,usubject):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = usubject
    msg['From'] = 'report@talentsumo.co'
    msg['To'] = uemail
    part2 = MIMEText(umessage, 'html')
    msg.attach(part2)
    s = smtplib.SMTP('smtp.acumbamail.com',int(587))
    s.login('deb@talentsumo.co', '9sHNPoWF7dXBWfrfdtmQ')
    s.sendmail('report@talentsumo.co', uemail, msg.as_string())
    s.quit()


def connectionTest(request):
    message = ''
    try:
        connection = mysql.connector.connect(host='',
                                         database='techtvxs_interview',
                                         user='techtvxs_phpinterview',
                                         password='(%G3CBBg}pBg')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL Server version ", db_Info)
            cursor = connection.cursor()
            cursor.callproc('getcolmnData1', ['ans_tmr8v3h9', ])
            people = ''
            for result in cursor.stored_results():
                people=result.fetchall()
                
            for p in people:
                print(p)

            cursor.execute("select database();")
            record = cursor.fetchone()
            message = "You're connected to database: " + ''.join(record)
            print("You're connected to database: ", record)
            cursor.close()
            connection.close()
    except Error as e:
            message = "Error while connecting to MySQL" + e
            print("Error while connecting to MySQL", e)
    return HttpResponse(message)
    

def activate(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = RegistrationTbl.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, RegistrationTbl.DoesNotExist):
        user = None
    if user is not None and account_activation_token.check_token(user, token):
        if(user.is_active != True):
            user.is_active = True
            user.save()
            connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
            if connection.is_connected():
                cursor1 = connection.cursor()
                cursor1.execute("SELECT COUNT(*) FROM `registration_tbl` WHERE email="+'"'+user.email+'"')
                rowcount = cursor1.fetchone()
                if rowcount[0]==0:  
                    stmt1 = "INSERT INTO `registration_tbl`(`fullname`, `email`, `password`) VALUES ('"+user.fullname+"','"+user.email+"','"+user.password+"');"
                    cursor1.execute(stmt1)
                    connection.commit()
                else:
                    stmt1 = "UPDATE `registration_tbl` SET `fullname`='"+user.fullname+"',`password`='"+user.password+"' WHERE email='"+user.email+"';"
                    cursor1.execute(stmt1)
                    connection.commit()
                getactivateUser = "SELECT id FROM `registration_tbl` WHERE email='"+user.email+"';"
                cursor1.execute(getactivateUser)
                userId = cursor1.fetchone()
                cursor1.execute("INSERT INTO `foldername`(`folderName`,`user_Id`) VALUES ('"+str("Default Folder")+"','"+str(userId[0])+"')")
                connection.commit()
                cursor1.close()
            return render(request , 'login.html')
        else:
            return render(request , 'login.html')
    else:
        return render(request , 'login.html')


def reset(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = RegistrationTbl.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, RegistrationTbl.DoesNotExist):
        user = None
    if user is not None and account_activation_token.check_token(user, token):

        return render(request , 'reset_password.html',{"email":str(user.email)})
    else:
        return render(request , 'login.html')

@csrf_exempt
def member_Create(request):
    if request.method == 'POST':
        uemail = request.POST.get("m_Email")
        userId = request.session.get('userId')
        userdbId = request.session.get('dbId')
        userEmail = request.session.get('userEmail')
        userName = request.session.get('userName')
        userDetails = RegistrationTbl.objects.get(id=userId)
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        if RegistrationTbl.objects.filter(email=uemail).exists():
            mreguserDetails = RegistrationTbl.objects.get(email=uemail)
            muserDetails = MemberMapping(member_of = userDetails,user=mreguserDetails)
            muserDetails.save()
            cursor1 = connection.cursor()
            stmt1 = "SELECT `id` from  `registration_tbl` where email='"+uemail+"';"
            cursor1.execute(stmt1)
            rows1 = cursor1.fetchone()
            mid = rows1[0]
            stmt2 = "INSERT INTO `tbl_member_mapping`(`member_of`,`userid`) VALUES ('"+str(userdbId)+"','"+str(mid)+"');"
            cursor1.execute(stmt2)
            connection.commit()
            cursor1.close()
            sendMemberNotification(userName,uemail)
        else:
            RuserDetails = RegistrationTbl(fullname='',email=uemail,password='',is_active=0)
            RuserDetails.save()
            muserDetails = MemberMapping(member_of=userDetails,user=RuserDetails)
            muserDetails.save()
            cursor1 = connection.cursor()
            stmt1 = "INSERT INTO `registration_tbl`(`email`) VALUES ('"+uemail+"');"
            cursor1.execute(stmt1)
            connection.commit()
            stmt2 = "INSERT INTO `tbl_member_mapping`(`member_of`,`userid`) VALUES ('"+str(userdbId)+"','"+str(cursor1.lastrowid)+"');"
            cursor1.execute(stmt2)
            connection.commit()
            cursor1.close()
            current_site = get_current_site(request)
            url = "http://" + current_site.domain + "/registration/" + str(urlsafe_base64_encode(force_bytes(userDetails.pk))) + '/' + str(account_activation_token.make_token(userDetails))
            sendRegistrationLink(userName,uemail,url)
        return HttpResponse('success')


def activatergst(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = RegistrationTbl.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, RegistrationTbl.DoesNotExist):
        user = None
    if user is not None and account_activation_token.check_token(user, token):
        print(user.email)
        return render(request , 'registration.html',{"email":str(user.email),"id":str(user.pk)})
    else:
        return HttpResponse('Activation link is invalid!')

@csrf_exempt
def shareformView(request):
    if request.method == 'POST':
        formId = request.POST.get("formId")
        formLongId = request.POST.get("formLongId")
        userId = request.session.get('dbId')
        userList = request.POST['userList']
        data = json.loads(userList)
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        atblName = "ans_"+str(formLongId[-8:]).lower()
        for l in data:
           stmt =  "SELECT * FROM `shared_table` where form_id='"+str(formId)+"' and shared_to='"+str(l)+"' and shared_by='"+str(userId)+"';"
           cursor.execute(stmt)
           rowcount = cursor.fetchone()
           if rowcount==None:  
               stmt1 = "INSERT INTO `shared_table`( `form_id`, `table_name`, `shared_by`, `shared_to`) VALUES ('"+str(formId)+"','"+str(atblName)+"','"+str(userId)+"','"+str(l)+"');"
               cursor.execute(stmt1)
               connection.commit()
        cursor.close()
        return HttpResponse("success")

@csrf_exempt
def deleteMemberView(request):
    if request.method == 'POST':
        memId = request.POST.get("memId")
        memEmail = request.POST.get("mEmail")
        userId = request.session.get('userId')
        userdbId = request.session.get('dbId')
        muserDetails = RegistrationTbl.objects.get(email=memEmail)
        userDetails = RegistrationTbl.objects.get(pk=userId)
        MemberMapping.objects.get(member_of=userDetails,user=muserDetails).delete()
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        stmt1 = " DELETE FROM `tbl_member_mapping` WHERE member_of='"+str(userdbId)+"' and userid = '"+str(memId)+"';"
        cursor.execute(stmt1)
        connection.commit()
        cursor.close()
        return HttpResponse("success")


@csrf_exempt
def chksmtp_settingsview(request):
    if request.method == 'POST':
        result= ''
        uformID = request.POST.get("formId")
        uformType = request.POST.get("formType")
        data = []
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        userdbId = request.session.get('dbId')
        getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid = '"+str(userdbId)+"';"
        cursor.execute(getsettings)
        results = cursor.fetchall()
        if(cursor.rowcount == 0):
            getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid  = 51;"
            cursor.execute(getsettings)
            results = cursor.fetchall()
        if(results[5][1]=='true' and results[9][1]=='true' and results[11][1]=='true' and results[7][1]=='true' and results[13][1]=='true'):
            result = 'Valid'
        else:
            result = 'Not Valid'
        if(uformType == "quiz" or uformType == "interview" or uformType == "viva"):
            getsettings= "SELECT `answer_link` FROM forms_templates WHERE `id` = '"+str(uformID)+"';"
            cursor.execute(getsettings)
            row = cursor.fetchone()
            if(str(row[0]).strip() != 'None'):
                ans_result = "Valid"
            else:
                ans_result = "Not Valid"
        else:
            ans_result = "Not Valid"
        response = {"smtp": result,"answerupload": ans_result}
        cursor.close()
        return JsonResponse(response,safe=False)

@csrf_exempt
def check_code_view(request):
    if request.method == 'POST':
        result = ''
        formId = request.POST.get("formId")
        formName = request.POST.get("formName")
        formCode = request.POST.get("formCode")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        gettblName = "SELECT `tbl_Name` FROM `forms_templates` WHERE `form_ID`='"+str(formId)+"' and `formlock_code`='"+str(formCode)+"' "
        cursor.execute(gettblName)
        tblName = cursor.fetchone()
        if(tblName != None):
            result = 'success'
        else:
            result = 'not valid'
        return HttpResponse(result)
@csrf_exempt
def get_responseCount_View(request):
    if request.method == 'POST':
        rcount = ''
        tdata = ''
        negative_marking = ''
        score_settings = ''
        userplan = ''
        try:
            formId = request.POST.get("formId")
            formName = request.POST.get("formName")
            connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
            cursor = connection.cursor()
            gettblName = "SELECT `ft`.`tbl_Name`,`ft`.`id`,`rt`.`email`,`ft`.`negative_marking`,`ft`.`score_settings` FROM `forms_templates` as ft JOIN `registration_tbl` as rt on `rt`.`id` = `ft`.`user_Id` WHERE `form_ID`='"+str(formId)+"'"
            cursor.execute(gettblName)
            tblName = cursor.fetchone()
            triggerData = "SELECT `fieldData` FROM `formfields` WHERE `formid`='"+str(tblName[1])+"' and fieldName='trigger'"
            cursor.execute(triggerData)
            trigger = cursor.fetchone()
            tdata = str(trigger[0])
            negative_marking = str(tblName[3])
            score_settings = str(tblName[4])
            useremail = str(tblName[2])
            userDetails = RegistrationTbl.objects.get(email = useremail)
            responseCount = "SELECT COUNT(*) FROM `"+str(tblName[0])+"`"
            cursor.execute(responseCount)
            resCount = cursor.fetchone()
            rcount = str(resCount[0])
            cursor.close()
            userplan = userDetails.Plan.name
        except:
            rcount = 0
            formId = request.POST.get("formId")
            formName = request.POST.get("formName")
            connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
            cursor = connection.cursor()
            gettblName = "SELECT `ft`.`tbl_Name`,`ft`.`id`,`rt`.`email`,`ft`.`negative_marking`,`ft`.`score_settings` FROM `forms_templates` as ft JOIN `registration_tbl` as rt on `rt`.`id` = `ft`.`user_Id` WHERE `form_ID`='"+str(formId)+"'"
            cursor.execute(gettblName)
            tblName = cursor.fetchone()
            negative_marking = str(tblName[3])
            score_settings = str(tblName[4])
            useremail = str(tblName[2])
            userDetails = RegistrationTbl.objects.get(email = useremail)
            userplan = userDetails.Plan.name
            triggerData = "SELECT `fieldData` FROM `formfields` WHERE formid ='"+str(tblName[1])+"' and fieldName='trigger'"
            cursor.execute(triggerData)
            trigger = cursor.fetchone()
            if(trigger != None):
                tdata = str(trigger[0])
            else:
                tdata = ''
        response = {'count':rcount,'trigger':tdata,'marking':negative_marking,'settings':score_settings,'plan':userplan}
        return JsonResponse(response,safe=False)
    elif request.method == 'GET':
        return HttpResponse(200)

def get_formCount_view(request):
    if request.method == 'GET':
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        formCount = "SELECT COUNT(*) FROM `forms_templates` where `user_Id` = '"+str(userdbId)+"' and `Duplicate` IS NULL;"        
        cursor.execute(formCount)
        formCountval = cursor.fetchone()
        fcount = str(formCountval[0])
        cursor.close()
        return JsonResponse(fcount,safe=False)

def get_membercount_view(request):
    if request.method == 'GET':
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        memberCount = "SELECT COUNT(*) FROM `tbl_member_mapping` where `member_of` = '"+str(userdbId)+"'"
        cursor.execute(memberCount)
        mebCount = cursor.fetchone()
        mcount = str(mebCount[0])
        cursor.close()
        return JsonResponse(mcount,safe=False)
        
@csrf_exempt
@require_http_methods(["POST"])
#@require_POST
def fileupload_View(request):
    files = request.FILES.getlist('file')
    url = ''
    storage_client = storage.Client.from_service_account_json('InterviewApp/interviewapp.json')
    bucket_name = "tal_interview_app_files"
    bucket = storage_client.get_bucket(bucket_name)
    if len(files) != 0:
        for file in files:
            blob = bucket.blob(str(file.name.replace(' ','_')))
            blob.upload_from_file(file)
            #fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Files/')
            #file_path=fs.save(file.name.replace(' ','_'),file) 
            #url += '/media/Forms/'+file_path
            blob.make_public()
            url += blob.public_url
    else:
        print('No File') 
    return HttpResponse(url)    
    
class videoindexView(View):
	return_url = None
	def get(self , request,Id=0):
		return render(request , 'videorecording.html')

class test_View(View):
    return_url = None
    def get(self , request,Id=0):
        return render(request , 'test1.html')
		
class videostatuserror(View):
	return_url = None
	def get(self , request,id=0):
		return render(request , 'servereerror.html')

class videointerneterror(View):
	return_url = None
	def get(self , request,id=0):
		return render(request , 'nointernet.html')
		
class videopreviewView(View):
	return_url = None
	def get(self , request,id=0):
		return render(request , 'videopreview.html')

def is_exist(object):
    client = storage.Client.from_service_account_json('/home/techtvxs/Interview/InterviewApp/interviewapp.json')
    bucket = client.bucket("tal_interview_app_videos")
    blob = bucket.get_blob(object)
    try:
        return blob.exists(client)
    except:
        return False

@csrf_exempt   
def checkvideoview(request):
    if request.method =="POST":
        fileName = request.POST.get('uFileName')
        filestatus = ''
        if(is_exist(fileName)== False):
            filestatus = 'not-exits'
        else:
            filestatus = 'exits'
        return HttpResponse(filestatus) 
        
    
@csrf_exempt
def video_InsertVideo(request):
    if request.method == 'POST':  
        file =  request.FILES['file']
        fileName= request.POST['filename']
        existingPath = request.POST['existingPath']
        end = request.POST['end']
        nextSlice = request.POST['nextSlice']
        intcounter = request.POST['counter']
        if file==""  or existingPath=="" or end=="":
            res = JsonResponse({'data':'Invalid Request'})
            return res
        else:
            if existingPath == 'null':
                if int(end):
                    p = videoProcess(file,fileName)
                    p.start()
                    res = JsonResponse({'data':'Uploaded Successfully','existingPath': fileName})
                else:
                    with open(fileName.replace('.mp4','')+intcounter+".txt", "wb") as handle:
                        handle.write(file.read())
                    res = JsonResponse({'data':'Uploaded Successfully','existingPath': fileName})
                return res
            else:
                if int(end):
                    with open(existingPath.replace('.mp4','')+intcounter+".txt", "wb") as handle:
                        handle.write(file.read())
                    p = videoProcess1(existingPath,intcounter)
                    p.start()
                    res = JsonResponse({'data':'Uploaded Successfully','existingPath':existingPath})
                else:
                    with open(existingPath.replace('.mp4','')+intcounter+".txt", "wb") as handle:
                        handle.write(file.read())
                    res = JsonResponse({'existingPath':existingPath})
                return res
                
@csrf_exempt
def video_Record_InsertVideo(request):
    if request.method == 'POST':  
        file =  request.FILES['file']
        existingPath = request.POST.get("existingPath")
        end =  request.POST.get("end")
        intcounter =request.POST.get("counter")
        if file==""  or existingPath=="" or end=="":
            res = JsonResponse({'data':'Invalid Request'})
            return res
        else:
            if existingPath == 'null':
                fileName = str(uuid.uuid4())+".mp4"
                if int(end):
                    p = videoProcess(file,fileName)
                    p.start()
                    res = JsonResponse({'data':'Uploaded Successfully','existingPath': fileName})
                else:
                    with open(fileName.replace('.mp4','')+intcounter+".txt", "wb") as handle:
                        handle.write(file.read())
                    res = JsonResponse({'data':'Uploaded Successfully','existingPath': fileName})
                return res
            else:
                if int(end):
                    with open(existingPath.replace('.mp4','')+intcounter+".txt", "wb") as handle:
                        handle.write(file.read())
                    p = videoProcess1(existingPath,intcounter)
                    p.start()
                    res = JsonResponse({'data':'Uploaded Successfully','existingPath':existingPath})
                else:
                    with open(existingPath.replace('.mp4','')+intcounter+".txt", "wb") as handle:
                        handle.write(file.read())
                    res = JsonResponse({'existingPath':existingPath})
                return res

@csrf_exempt
def add_form_lockView(request):
    if request.method == 'POST':  
        formId = request.POST.get("formId")
        lockCode =  request.POST.get("formlockCode")
        lockStatus =  request.POST.get("formlockstatus")
        lockPosition =  request.POST.get("formlockposition")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        updateformlockDetails = "UPDATE `forms_templates` SET `formlock_code`='"+lockCode+"',`form_lock_active`='"+lockStatus+"',`form_lock_position`='"+lockPosition+"' WHERE `id`='"+formId+"';"
        cursor.execute(updateformlockDetails)
        connection.commit()
        return HttpResponse('sucess') 


def check_list_view(request):
       if request.method == 'GET':  
        ulistName = request.GET['ulistName']
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("SELECT * from `tbl_students_lists` WHERE `user_Id` ='"+str(userdbId)+"' AND `list_name`= '"+str(ulistName)+"';")
        rowcount = cursor.fetchone()
        cursor.close()
        if rowcount != None:
            result= "Exits"
        else:
            result= "Valid"
        return HttpResponse(result) 

@csrf_exempt
def add_list_view(request):
    if request.method == 'POST':
        ulistName = request.POST.get("ulistName")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("INSERT INTO `tbl_students_lists`( `list_name`, `user_Id`) VALUES ('"+str(ulistName)+"','"+str(userdbId)+"')")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')

@csrf_exempt
def update_list_view(request):
    if request.method == 'POST':
        ulistId = request.POST.get("ulistid")
        ulistName = request.POST.get("ulistName")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("UPDATE `tbl_students_lists` SET `list_name`='"+str(ulistName)+"' WHERE user_Id ='"+str(userdbId)+"' AND `id`= '"+str(ulistId)+"'  ")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')

@csrf_exempt
def delete_list_view(request):
    if request.method == 'POST':
        ulistId = request.POST.get("ulistid")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("DELETE FROM `tbl_students_lists`  WHERE user_Id ='"+str(userdbId)+"' AND `id`= '"+str(ulistId)+"'")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')

@csrf_exempt
def import_students_view(request):
    if request.method == 'POST':
        files = request.FILES.getlist("file") 
        ulistName = request.POST.get("listid")
        uextension = request.POST.get("extension")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("INSERT INTO `tbl_students_lists`( `list_name`, `user_Id`) VALUES ('"+str(ulistName)+"','"+str(userdbId)+"')")
        connection.commit()
        listid = cursor.lastrowid 
        if len(files) != 0:
            for file in files:
                if(uextension == 'xlsx'):
                    file_in_memory = file.read()
                    wrkbk = load_workbook(filename=BytesIO(file_in_memory))
                    sh = wrkbk.active
                    for i, row in enumerate(sh.rows):
                        cursor.execute("INSERT INTO `tbl_students_details`( `student_name`, `student_email`, `student_collage`, `student_specialization`, `student_course`, `student_year`, `student_list_id`, `user_id`, `student_active`) VALUES ('"+str(sh.cell(row = i+1, column = 1).value)+"','"+str(sh.cell(row = i+1, column = 2).value)+"','"+str(sh.cell(row = i+1, column = 3).value)+"','"+str(sh.cell(row = i+1, column = 4).value)+"','"+str(sh.cell(row = i+1, column = 5).value)+"','"+str(sh.cell(row = i+1, column = 6).value)+"','"+str(listid)+"','"+str(userdbId)+"','active');")
                        connection.commit()
                elif(uextension == 'xls'):
                    wb = xlrd.open_workbook(file_contents=file.read())
                    sheet = wb.sheet_by_index(0)
                    for i in range(1,sheet.nrows):
                        #print(sheet.row(i)[0].value)
                        cursor.execute("INSERT INTO `tbl_students_details`( `student_name`, `student_email`, `student_collage`, `student_specialization`, `student_course`, `student_year`, `student_list_id`, `user_id`, `student_active`) VALUES ('"+str(sheet.row(i)[0].value)+"','"+str(sheet.row(i)[1].value)+"','"+str(sheet.row(i)[2].value)+"','"+str(sheet.row(i)[3].value)+"','"+str(sheet.row(i)[4].value)+"','"+str(sheet.row(i)[5].value)+"','"+str(listid)+"','"+str(userdbId)+"','active');")
                        connection.commit()
                elif(uextension == 'csv'):
                    rowlists = []
                    reader = csv.DictReader(io.StringIO(file.read().decode('utf-8')))
                    for num, line in enumerate(reader): 
                        rowarray = []
                        for l in list(list(line.items())):
                            rowarray.append(l[1])
                        rowlists.append(rowarray)
                    for row in rowlists:
                        cursor.execute("INSERT INTO `tbl_students_details`( `student_name`, `student_email`, `student_collage`, `student_specialization`, `student_course`, `student_year`, `student_list_id`, `user_id`, `student_active`) VALUES ('"+str(row[0])+"','"+str(row[1])+"','"+str(row[2])+"','"+str(row[3])+"','"+str(row[4])+"','"+str(row[5])+"','"+str(listid)+"','"+str(userdbId)+"','active');")
                        connection.commit()
                connection.close()
                url = 'sucess'
        else:
            print('No File')
            url = 'No File'
        return HttpResponse(url)

def check_student_view(request):
       if request.method == 'GET':  
        ulistEmail = request.GET['ustudentEmail']
        ulistname = request.GET['ulistName']
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("SELECT `tbl_students_details`.* FROM `tbl_students_details` INNER JOIN `tbl_students_lists` ON `tbl_students_details`.student_list_id = `tbl_students_lists`.`id` WHERE `tbl_students_lists`.`user_id`='"+str(userdbId)+"' AND `tbl_students_lists`.list_name= '"+str(ulistname)+"' AND `tbl_students_details`.`student_email`='"+str(ulistEmail)+"' ORDER BY `tbl_students_lists`.`user_id`;")
        rowcount = cursor.fetchone()
        cursor.close()
        if rowcount != None:
            result= "Exits"
        else:
            result= "Valid"
        return HttpResponse(result) 

@csrf_exempt
def add_student_view(request):
    if request.method == 'POST':
        ulistName = request.POST.get("ulistName")
        ustudName = request.POST.get("ustudName")
        ustudEmail = request.POST.get("ustudEmail")
        ustudCourse = request.POST.get("ustudCourse")
        ustudSpecial = request.POST.get("ustudSpecial")
        ustudCollage = request.POST.get("ustudCollage")
        ustudyear = request.POST.get("ustudyear")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("SELECT id from tbl_students_lists where list_name = '"+str(ulistName)+"';")
        id = cursor.fetchone()
        cursor.execute("INSERT INTO `tbl_students_details`( `student_name`, `student_email`, `student_collage`, `student_specialization`, `student_course`, `student_year`, `student_list_id`, `user_id`, `student_active`) VALUES ('"+str(ustudName)+"','"+str(ustudEmail)+"','"+str(ustudCollage)+"','"+str(ustudSpecial)+"','"+str(ustudCourse)+"','"+str(ustudyear)+"','"+str(id[0])+"','"+str(userdbId)+"','active');")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')

@csrf_exempt
def update_student_view(request):
    if request.method == 'POST':
        ulistName = request.POST.get("ulistName")
        ustudId = request.POST.get("ustudId")
        ustudName = request.POST.get("ustudName")
        ustudEmail = request.POST.get("ustudEmail")
        ustudCourse = request.POST.get("ustudCourse")
        ustudSpecial = request.POST.get("ustudSpecial")
        ustudCollage = request.POST.get("ustudCollage")
        ustudyear = request.POST.get("ustudyear")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("UPDATE `tbl_students_details` SET `student_name`='"+str(ustudName)+"',`student_email`='"+str(ustudEmail)+"',`student_collage`='"+str(ustudCollage)+"',`student_specialization`='"+str(ustudSpecial)+"',`student_course`='"+str(ustudCourse)+"',`student_year`='"+str(ustudyear)+"' WHERE id='"+str(ustudId)+"' AND user_id='"+str(userdbId)+"';")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')

@csrf_exempt
def delete_student_view(request):
    if request.method == 'POST':
        ustudeId = request.POST.get("ustudid")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("DELETE FROM `tbl_students_details`  WHERE user_Id ='"+str(userdbId)+"' AND `id`= '"+str(ustudeId)+"'")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')

@csrf_exempt
def block_student_view(request):
    if request.method == 'POST':
        ustudId = request.POST.get("ustudid")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("UPDATE `tbl_students_details` SET `student_active`='Blocked' WHERE id='"+str(ustudId)+"' AND user_id='"+str(userdbId)+"';")
        connection.commit()
        rowcount = cursor.fetchone()
        cursor.close()
        return HttpResponse('success')
        
@csrf_exempt
def check_campaign_view(request):
       if request.method == 'POST':  
        ucamp_Name = request.POST.get("camp_name")
        ucamp_Date = request.POST.get("camp_date")
        ucamp_List = request.POST['camp_lists']
        ucamp_data = json.loads(ucamp_List)
        ucamp_List_Name = request.POST['camp_list_names']
        ucamp_data_names = json.loads(ucamp_List_Name)
        ucamp_formId = request.POST.get("camp_formId")
        userdbId = request.session.get('dbId')
        errorlist = []
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.execute("SELECT * from `tbl_campaign_details` WHERE `user_id` ='"+str(userdbId)+"' AND `campaign_Name`= '"+str(ucamp_Name)+"';")
        rowcount = cursor.fetchone()
        if rowcount != None:
            result= "Campaign Name Already exists."
        else:
            result= "Valid"
            for i,listid in enumerate(ucamp_data):
                cursor.execute("SELECT * from `tbl_campaign_details` WHERE `user_id` ='"+str(userdbId)+"' AND `campaign_list`= '"+str(listid)+"' AND `campaign_Date`= '"+str(ucamp_Date)+"';")
                rowcount = cursor.fetchone()
                if rowcount != None:
                    errorlist.append(ucamp_data_names[i])
                else:
                    pass
            if(len(errorlist) != 0):
                result = "Campaign already running for this "+ ', '.join([str(elem) for elem in errorlist])+" on selected date" 
            else:
                result= "Valid"  
        cursor.close()
        return HttpResponse(result)

@csrf_exempt
def add_campaign_view(request):
    if request.method == 'POST':
        ucamp_Name = request.POST.get("camp_name")
        ucamp_Date = request.POST.get("camp_date")
        ucamp_List = request.POST['camp_lists']
        ucamp_data = json.loads(ucamp_List)
        ucamp_formId = request.POST.get("camp_formId")
        userdbId = request.session.get('dbId')
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        for listid in ucamp_data:
            stmt1 = "INSERT INTO `tbl_campaign_details`(`campaign_Name`, `campaign_list`, `user_id`, `campaign_Date`, `campaign_Status`, `formid`) VALUES ('"+str(ucamp_Name)+"' ,'"+str(listid)+"','"+str(userdbId)+"' ,'"+ucamp_Date+"','Active','"+str(ucamp_formId)+"')"
            cursor.execute(stmt1)
            connection.commit()
        cursor.close()
        return HttpResponse('success')

def sendEmailCampaign(userid,camp_list,intervlink,type):
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid = '"+str(userid)+"';"
    cursor.execute(getsettings)
    results = cursor.fetchall()
    if(cursor.rowcount == 0):
        getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid  = 51;"
        cursor.execute(getsettings)
        results = cursor.fetchall()
    host = ''
    port = ''
    login = ''  
    password = ''
    fromaddr = '' 
    message = ''
    if(len(results) != 0):
        host = results[0][1]
        port = results[1][1]
        login = results[2][1]
        password = results[3][1]
        fromaddr =  results[4][1]
        if(type=="campMessage"):
            campMessgtatus = results[13][1]
            if(campMessgtatus == 'true'):
                cursor.execute("SELECT * from `tbl_students_details` WHERE  `student_list_id`= '"+str(camp_list)+"';")
                stu_details = cursor.fetchall()
                for student in stu_details:
                    list_studetails = list(student)
                    stud_name = list_studetails[1]
                    stud_email = list_studetails[2]
                    stud_collage = list_studetails[3]
                    stud_spec = list_studetails[4]
                    stud_course = list_studetails[5]
                    stud_year = list_studetails[6]
                    cMessgtext = results[14][1]
                    cMessgtext = str(cMessgtext)
                    cMessgtext = cMessgtext.replace('@Name', stud_name)
                    cMessgtext = cMessgtext.replace('@Email', stud_email)
                    cMessgtext = cMessgtext.replace('@Collage', stud_collage)
                    cMessgtext = cMessgtext.replace('@Specialization', stud_spec)
                    cMessgtext = cMessgtext.replace('@Course', stud_course)
                    cMessgtext = cMessgtext.replace('@Year', stud_year)
                    cMessgtext = cMessgtext.replace("@InterviewLink", intervlink)
                    mailstatus = SendEmailAct(fromaddr,stud_email,login,password,host,port,cMessgtext,'Career Interview Link')
                    print(mailstatus)
                cursor.close()
                return "sent"

class textCampaignmodule(View):
    return_url = None
    def get(self , request,Id=''):
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        today = date.today()
        cursor.execute("SELECT * from `tbl_campaign_details` WHERE  `campaign_Date`= '"+today.strftime("%Y/%m/%d")+"' and campaign_Status='Active';")
        rowcount = cursor.fetchall()
        if rowcount != None:
            for rows in rowcount:
                campdetails = list(rows)
                cursor.execute("SELECT `form_ID` from `forms_templates` WHERE  `id`= '"+str(campdetails[6])+"';")
                intervlink = 'https://spellcheck.techdivaa.com/'+cursor.fetchone()[0]
                user_id = str(campdetails[3])
                result = sendEmailCampaign(user_id,str(campdetails[2]),intervlink,'campMessage')
                if(result == 'sent'):
                    stmt1 = "UPDATE `tbl_campaign_details` SET `campaign_Status`='Completed' WHERE id='"+str(campdetails[0])+"'"
                    cursor.execute(stmt1)
                    connection.commit()
        cursor.close()
        return HttpResponse("sucess")


class campaignView(View):
    return_url = None
    def get(self , request,Id=''):
        userId = request.session.get('userId')
        if(userId!=None):
            return render(request , 'campaign.html')
        else:
            return redirect('login')


def get_CampaignData(request):
    if(request.method == 'GET'):
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        userdbId = request.session.get('dbId')
        getcampdetals= "SELECT c.id, c.campaign_Name,t.`form_Name`,s.`list_name`, COUNT(p.student_list_id) as stu_count,DATE_FORMAT(c.campaign_Date, '%d-%m-%Y') AS schedule_on,c.campaign_Status FROM `tbl_campaign_details` c LEFT JOIN `tbl_students_details` p ON c.campaign_list = p.student_list_id JOIN `forms_templates` t ON c.`formid` = t.`id` JOIN `tbl_students_lists` s ON c.`campaign_list` = s.`id` where  c.`user_id` = '"+str(userdbId)+"' GROUP BY c.id;"
        cursor.execute(getcampdetals)
        allrows = cursor.fetchall()
        cursor.close()
        return JsonResponse(allrows,safe=False)
        
def pause_campaign_View(request):
    if(request.method == 'GET'):
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        userdbId = request.session.get('dbId')
        campid = request.GET['uCampaignId']
        pausecapmdtls = "UPDATE `tbl_campaign_details` SET `campaign_Status`='Paused' WHERE `user_id`='"+str(userdbId)+"'  and `id`='"+str(campid)+"';"
        cursor.execute(pausecapmdtls)
        connection.commit()
        cursor.close()
        return HttpResponse('sucess')

def delete_campaign_View(request):
    if(request.method == 'GET'):
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        userdbId = request.session.get('dbId')
        campid = request.GET['uCampaignId']
        pausecapmdtls = "DELETE FROM `tbl_campaign_details` WHERE `user_id`='"+str(userdbId)+"'  and `id`='"+str(campid)+"';"
        cursor.execute(pausecapmdtls)
        connection.commit()
        cursor.close()
        return HttpResponse('sucess')



@csrf_exempt
def upload_importInterview_view(request):
    if request.method == 'POST':
        url = ''
        files = request.FILES.get("file") 
        uurl = request.POST.get("url")
        uformtype = request.POST.get("formtype")
        ugoogle = request.POST.get("google")
        ufileName = request.POST.get("fileName")
        userdbId = request.session.get('dbId')
        uextension = request.POST.get("extension")
        interviewcolumns = ["S.No .","Question","Answer Type","Answer","Key words"]
        ansfileName = 'ans_link_'+ufileName
        quizcolumns = ["S.No .","Question","Answer Type","Option 1","Option 2","Option 3","Option 4","Right Answer","Reason"]
        if(ugoogle == 'yes'):
            url_1 = uurl.replace("/edit#gid=", "/export?format=csv&gid=0").replace("/edit?usp=sharing", "/export?format=csv&gid=0")
            pddatframe = pd.read_csv(url_1, header=None)
            columnvalue = []
            arrayofrows = []
            columindex = []
            default_Answer = ''
            interviewName = str(pddatframe[5][0]).strip()
            for i in range(0,len(pddatframe.columns)):
                try:
                    columnvalue.append(str(pddatframe[i][3]).strip())
                    if('Option' in str(pddatframe[i][3]).strip()):
                        columindex.append(i)
                    elif('Right Answer' in str(pddatframe[i][3]).strip()):
                        default_Answer = i;
                except:
                    pass
            errorMessage = ''
            if(uformtype == 'interview'):
                if(all(item in columnvalue for item in interviewcolumns)):
                    errorMessage = "sucess"
                    pddatframe.to_excel(settings.MEDIA_ROOT +'/Interview_Templates/'+ansfileName, index=False, header=False)
                    file_path = 'media/Interview_Templates/'+ansfileName
                    for row in range(3,len(pddatframe)):
                        colomnlist = []
                        for i in range(0,len(pddatframe.columns)):
                            try:
                                if(str(pddatframe[i][row]).strip() != 'nan' or (i in columindex)):
                                        colomnlist.append(str(pddatframe[i][row]).strip())
                            except:
                                 pass
                        if(len(colomnlist) != 0):
                            arrayofrows.append(colomnlist)
                    importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                    return HttpResponse(importresult)
                else:
                    errorMessage = "not valid"
                    return HttpResponse(errorMessage)
            elif(uformtype == 'quiz'):
                if(all(item in columnvalue for item in quizcolumns)):
                    errorMessage = "sucess"
                    pddatframe.to_excel(settings.MEDIA_ROOT +'/Interview_Templates/'+ansfileName, index=False, header=False)
                    file_path = 'media/Interview_Templates/'+ansfileName
                    for row in range(3,len(pddatframe)):
                        colomnlist = []
                        for i in range(0,len(pddatframe.columns)):
                            try:
                                if(str(pddatframe[i][row]).strip() != 'nan' or (i in columindex)):
                                    colomnlist.append(str(pddatframe[i][row]).strip())
                            except:
                                pass
                        if(len(colomnlist) != 0):
                            arrayofrows.append(colomnlist)
                    importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                    return HttpResponse(importresult)
                else:
                    errorMessage = "not valid"
                    return HttpResponse(errorMessage)
        else:
            if(uextension == 'xlsx'):
                pddatframe = pd.read_excel(files.read(), header=None)
                columnvalue = []
                arrayofrows = []
                columindex = []
                default_Answer = ''
                interviewName = str(pddatframe[5][0]).strip()
                for i in range(0,len(pddatframe.columns)):
                    try:
                        columnvalue.append(str(pddatframe[i][3]).strip())
                        if('Option' in str(pddatframe[i][3]).strip()):
                            columindex.append(i)
                        elif('Right Answer' in str(pddatframe[i][3]).strip()):
                            default_Answer = i;
                    except:
                        pass
                errorMessage = ''
                if(uformtype == 'interview'):
                    if(all(item in columnvalue for item in interviewcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(pddatframe)):
                            colomnlist = []
                            for i in range(0,len(pddatframe.columns)):
                                try:
                                    if(str(pddatframe[i][row]).strip() != 'nan' or (i in columindex)):
                                        colomnlist.append(str(pddatframe[i][row]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                        return HttpResponse(importresult)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                elif(uformtype == 'quiz'):
                    if(all(item in columnvalue for item in quizcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(pddatframe)):
                            colomnlist = []
                            for i in range(0,len(pddatframe.columns)):
                                try:
                                    if(str(pddatframe[i][row]).strip() != 'nan' or (i in columindex)):
                                        colomnlist.append(str(pddatframe[i][row]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                        return HttpResponse(importresult)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
            elif(uextension == 'xls'):
                wb = xlrd.open_workbook(file_contents=files.read())
                sheet = wb.sheet_by_index(0)
                columnvalue = []
                arrayofrows = []
                columindex = []
                default_Answer = ''
                interviewName = str(sheet.row(0)[5].value).strip()
                for i in range(0,sheet.ncols):
                    try:
                        if(len(str(sheet.row(3)[i].value).strip()) != 0):
                            columnvalue.append(str(sheet.row(3)[i].value).strip())
                            if('Option' in str(sheet.row(3)[i].value).strip()):
                                    columindex.append(i)
                            elif('Right Answer' in str(sheet.row(3)[i].value).strip()):
                                default_Answer = i;
                    except:
                        pass
                if(uformtype == 'interview'):
                    if(all(item in columnvalue for item in interviewcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,sheet.nrows):
                            colomnlist = []
                            for i in range(0,sheet.ncols):
                                try:
                                    if(len(str(sheet.row(row)[i].value).strip()) != 0):
                                        colomnlist.append(str(sheet.row(row)[i].value).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                        return HttpResponse(importresult)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                elif(uformtype == 'quiz'):
                    if(all(item in columnvalue for item in quizcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,sheet.nrows):
                            colomnlist = []
                            for i in range(0,sheet.ncols):
                                try:
                                    if(len(str(sheet.row(row)[1].value).strip()) != 0):
                                        if(len(str(sheet.row(row)[i].value).strip()) != 0 or (i in columindex)):
                                            colomnlist.append(str(sheet.row(row)[i].value).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                        return HttpResponse(importresult)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                return HttpResponse('sucess')
            elif(uextension == 'csv'):
                reader = csv.reader(codecs.iterdecode(files, 'utf-8', errors='ignore'))
                csvlist = list(reader)
                columnvalue = []
                arrayofrows = []
                columindex = []
                default_Answer = ''
                interviewName = str(csvlist[0][5]).strip()
                for i in range(0,len(csvlist[3])):
                    try:
                        if(len(str(csvlist[3][i]).strip()) != 0):
                            columnvalue.append(str(csvlist[3][i]).strip())
                            if('Option' in str(csvlist[3][i]).strip()):
                                columindex.append(i)
                            elif('Right Answer' in str(csvlist[3][i]).strip()):
                                default_Answer = i;
                    except:
                        pass
                if(uformtype == 'interview'):
                    if(all(item in columnvalue for item in interviewcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(csvlist)):
                            colomnlist = []
                            for i in range(0,len(csvlist[row])):
                                try:
                                    if(len(str(csvlist[row][i]).strip()) != 0):
                                        colomnlist.append(str(csvlist[row][i]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                        return HttpResponse(importresult)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                elif(uformtype == 'quiz'):
                    if(all(item in columnvalue for item in quizcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(csvlist)):
                            colomnlist = []
                            for i in range(0,len(csvlist[row])):
                                try:
                                    if(len(str(csvlist[row][1]).strip()) != 0):
                                        if(len(str(csvlist[row][i]).strip()) != 0 or (i in columindex)):
                                            colomnlist.append(str(csvlist[row][i]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        importresult = functioncreateimportfrmview(interviewName,arrayofrows,uformtype,userdbId,columindex,file_path,default_Answer)
                        return HttpResponse(importresult)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                return HttpResponse('sucess')

def functioncreateimportfrmview(interviewName, arraylist, interviewtype,userid,colmnindex,ans_file_path,default_answ): 
    formid = ''
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    id = generateId()
    result = urllib.request.urlretrieve('https://i.imgur.com/1AGR1Q8.png')
    fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Forms/')
    fileName = "new_"+interviewName.replace(' ','_')
    fileName = fileName+'.png'
    file_path=fs.save(fileName,File(open(result[0], 'rb')))
    url = '/media/Forms/'+file_path
    atblName = "ans_"+str(id[-8:]).lower()
    cursor.execute("SELECT `id` FROM `foldername` WHERE `user_Id`='"+str(userid)+"'")
    rows = cursor.fetchall()
    folderid = ', '.join([str(elem) for elem in list(rows[0])])
    cursor.execute("INSERT INTO `forms_templates`( `form_Name`, `form_ID`,`tbl_Name`, `Parent_Template`, `form_img`, `formtype`,`user_Id`,`folder_id`,`answer_link`,`public`,`negative_marking`,`form_Question_type`,`import_mode`,`auto_report`) VALUES ('"+interviewName+"','"+id+"','"+atblName+"','Blank','"+url+"','form','"+str(userid)+"','"+str(folderid)+"','"+str(ans_file_path)+"','0','no','"+str(interviewtype)+"','import','no')")
    connection.commit()  
    formid = cursor.lastrowid 
    cursor7 = connection.cursor()
    cursor7.execute("CREATE TABLE "+atblName+" ("+atblName+"id INT(6) UNSIGNED AUTO_INCREMENT PRIMARY KEY, SubDate Text NULL,ipAddress Text NULL,commReview Text NULL,videoColloboration Text NULL,reviewStatus Text NULL,overGrade Text NULL,sent Text NULL,sentDate Text NULL,sentReport Text NULL,ReportLink Text NULL,AutoGrade Text NULL );")
    connection.commit()
    cursor7.execute("INSERT INTO "+atblName+"(`SubDate`,`ipAddress`,`commReview`,`videoColloboration`,`reviewStatus`,`overGrade`,`sent`,`sentDate`,`ReportLink`,`AutoGrade`) VALUES ('','','','','','','','','','');")
    connection.commit()
    cursor7.execute('INSERT INTO `form_column_mapping`( `formid`, `tblName`, `name`, `email`, `review`, `nominee`) VALUES ("'+id+'","'+atblName+'","","","","");')
    connection.commit()
    cursor7.close()
    lengtharray = len(arraylist)
    if(interviewtype == 'interview'):
        for i, array in enumerate(arraylist):
            htmltxt = ''
            listhtml = ''
            idnextval = ''
            textindex = 0
            typeindex = 0
            if(i == 1):
                prevval = 'step-0'
                idnextval = 'fieldques'+str(int(i+1))
            else:
                k = (int(i)-1)
                prevval = 'fieldques'+str(k)
                idnextval = 'fieldques'+str(int(i+1))
            idval ='fieldques'+str(int(i+1));
            elidval ='ques'+str(int(i+1));
            if(arraylist[i][1]=='Text'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid = 'shortText';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@shortText'+lbid+'" data-column="field'+idval+'" data-fieldValue="" data-v-a="false" data-formlock="false"><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe></div><input class="form-control input" type="text" data-type="text" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)" ><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-i-cursor"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@shortText'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false"><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe></div><input class="form-control input" type="text" data-type="text" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)" ><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-i-cursor"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
            elif(arraylist[i][1]=='Email'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid = 'emailAddress';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@emailAddress'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input class="form-control input" type="email" name="email" data-type="email" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)"><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="far fa-envelope"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@emailAddress'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input class="form-control input" type="email" name="email" data-type="email" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)"><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="far fa-envelope"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
            elif(arraylist[i][1]=='Video'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid = 'videoUpload';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@videoUpload'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input type="text" class="videourl" data-type="link" value=""><a class="videoclck" onclick="videoclick(this);"><i class="fas fa-video"></i></a><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-video"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@videoUpload'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input type="text" class="videourl" data-type="link" value=""><a class="videoclck" onclick="videoclick(this);"><i class="fas fa-video"></i></a><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button></p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-video"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
            elif(arraylist[i][1]=='Audio'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid = 'audioUpload';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@audioUpload'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input type="text" class="audiourl" data-type="link" value=""><a class="audioclck" onclick="audioclick(this);"><i class="fas fa-microphone-alt"></i></a><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-volume-up"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@audioUpload'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input type="text" class="audiourl" data-type="link" value=""><a class="audioclck" onclick="audioclick(this);"><i class="fas fa-microphone-alt"></i></a><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-volume-up"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
            cursor.execute("INSERT INTO `formfields`(`fieldName`, `fieldData`, `fieldsetData`, `formid`) VALUES ('"+str(elidval)+"','"+str(listhtml)+"','"+str(htmltxt)+"','"+str(formid)+"')")
            connection.commit()
    elif(interviewtype == 'quiz'):
        for i, array in enumerate(arraylist):
            coloptionsindex = []
            htmltxt = ''
            listhtml = ''
            idnextval = ''
            if(i == 1):
                prevval = 'step-0'
                idnextval = 'fieldques'+str(int(i+1))
            else:
                k = (int(i)-1)
                prevval = 'fieldques'+str(k)
                idnextval = 'fieldques'+str(int(i+1))
            idval ='fieldques'+str(int(i+1));
            elidval ='ques'+str(int(i+1));
            if(arraylist[i][1]=='Text'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid = 'shortText';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@shortText'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false"><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe></div><input class="form-control input" type="text" data-type="text" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)" ><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-i-cursor"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@shortText'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false"><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe></div><input class="form-control input" type="text" data-type="text" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)" ><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-i-cursor"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
            elif(arraylist[i][1]=='Email'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid ='emailAddress';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@emailAddress'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input class="form-control input" type="email" name="email" data-type="email" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)"><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="far fa-envelope"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@emailAddress'+lbid+'" data-column="field'+idval+'"  data-fieldValue="" data-v-a="false" data-formlock="false" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><input class="form-control input" type="email" name="email" data-type="email" onchange="inputonChange(this)" onkeypress="inputonkeypress(this)"><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="far fa-envelope"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>'
            elif(arraylist[i][1]=='MCQ'):
                choices = ''
                defaultAnswer = arraylist[i][int(default_answ)-1];
                defaultIndex = 0;
                choiceid= 'multich'+str(i)
                for cindex in range(0,len(colmnindex)-1):
                    if(len(arraylist[i][int(colmnindex[cindex]-1)])!=0):
                        if((arraylist[i][int(colmnindex[cindex]-1)] != 'None') and (arraylist[i][int(colmnindex[cindex]-1)] != 'nan')):
                            if(arraylist[i][int(colmnindex[cindex]-1)] == defaultAnswer):
                                defaultIndex = cindex
                            choices += '<div class="p-2 rounded checkbox-form"><div class="form-check"> <input class="form-check-input" type="checkbox" data-type="choice" name="'+choiceid+'" id="'+choiceid+str((cindex+1))+'" value="'+arraylist[i][int(colmnindex[cindex]-1)]+'" onclick="inputonChange(this)" onkeypress="inputonkeypress(this)"> <label class=" form-check-label" for="'+choiceid+str((cindex+1))+'""> '+arraylist[i][int(colmnindex[cindex]-1)]+' </label> </div></div>'
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid ='multipleChoice';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@multipleChoice'+lbid+'" data-column="field'+idval+'"  default-val="" data-fieldValue="'+arraylist[i][int(default_answ)-1]+'" default-val="'+arraylist[i][int(default_answ)-1]+'" data-v-a="false" data-formlock="false" data-scoreobject=""><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="checkbox-group" data-assign="false" data-default="'+str(defaultIndex)+'" data-reason="">'+choices+'</div><div class="supportlbl"><input type="hidden" class="multiplehidden" value=""><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-check-square"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@multipleChoice'+lbid+'" data-column="field'+idval+'"  default-val="" data-fieldValue="'+arraylist[i][int(default_answ)-1]+'" default-val="'+arraylist[i][int(default_answ)-1]+'" data-v-a="false" data-formlock="false" data-scoreobject=""><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="checkbox-group" data-assign="false" data-default="'+str(defaultIndex)+'" data-reason="">'+choices+'</div><div class="supportlbl"><input type="hidden" class="multiplehidden" value=""><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button></p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-check-square"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>'
            elif(arraylist[i][1]=='Dropdown'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                choices = ''
                defaultAnswer = arraylist[i][int(default_answ)-1];
                defaultIndex = 0;
                for cindex in range(0,len(colmnindex)-1):
                    if(len(arraylist[i][int(colmnindex[cindex]-1)])!=0):
                        if((arraylist[i][int(colmnindex[cindex]-1)] != 'None') and (arraylist[i][int(colmnindex[cindex]-1)] != 'nan')):
                            if(arraylist[i][int(colmnindex[cindex]-1)] == defaultAnswer):
                                defaultIndex = cindex
                            choices += '<option value="'+arraylist[i][int(colmnindex[cindex]-1)]+'">'+arraylist[i][int(colmnindex[cindex]-1)]+'</option> '
                lbid ='dropdownChoice';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@dropdownChoice'+lbid+'" data-column="field'+idval+'"  data-fieldValue="'+arraylist[i][int(default_answ)-1]+'" default-val="'+arraylist[i][int(default_answ)-1]+'"  default-val="" data-v-a="false" data-formlock="false" data-scoreobject=""><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="selectdiv"><label><select class="form-control input" onchange="inputonChange(this)" data-type="select" data-assign="false" data-default="'+str(defaultIndex)+'" data-reason="" onkeypress="inputonkeypress(this)">'+choices+'  </select> </label></div><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-chevron-circle-down"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@dropdownChoice'+lbid+'" data-column="field'+idval+'"  data-fieldValue="'+arraylist[i][int(default_answ)-1]+'" default-val="'+arraylist[i][int(default_answ)-1]+'"  default-val="" data-v-a="false" data-formlock="false" data-scoreobject=""><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="selectdiv"><label><select class="form-control input" onchange="inputonChange(this)" data-type="select" data-assign="false" data-default="'+str(defaultIndex)+'" data-reason="" onkeypress="inputonkeypress(this)"> '+choices+'  </select> </label></div><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-chevron-circle-down"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>'
            elif(arraylist[i][1]=='Scale'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                choices = ''
                defaultAnswer = arraylist[i][int(default_answ)-1];
                defaultIndex = 0;
                for cindex in range(0,len(colmnindex)-1):
                    if(len(arraylist[i][int(colmnindex[cindex]-1)])!=0):
                        if((arraylist[i][int(colmnindex[cindex]-1)] != 'None') and (arraylist[i][int(colmnindex[cindex]-1)] != 'nan')):
                            if(arraylist[i][int(colmnindex[cindex]-1)] == defaultAnswer):
                                defaultIndex = cindex
                            choices += '<option value="'+arraylist[i][int(colmnindex[cindex]-1)]+'">'+arraylist[i][int(colmnindex[cindex]-1)]+'</option> '
                lbid ='reviewChoice';
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@reviewChoice'+lbid+'" data-fieldValue="'+arraylist[i][int(default_answ)-1]+'"  data-column="field'+idval+'"  default-val="'+arraylist[i][int(default_answ)-1]+'"  data-v-a="false" data-formlock="false" data-scoreobject="" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="box box-orange box-example-1to10"><div class="box-body"> <select class="rating-scale" name="rating" data-type="select" autocomplete="off">'+choices+'</select> </div></div><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fab fa-cloudscale"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@reviewChoice'+lbid+'" data-fieldValue="'+arraylist[i][int(default_answ)-1]+'" data-column="field'+idval+'"  default-val="'+arraylist[i][int(default_answ)-1]+'" data-v-a="false" data-formlock="false" data-scoreobject="" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="box box-orange box-example-1to10"><div class="box-body"> <select class="rating-scale" name="rating" data-type="select" autocomplete="off">'+choices+' </select> </div></div><div class="supportlbl"><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button></p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fab fa-cloudscale"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>'
            elif(arraylist[i][1]=='True/False'):
                idText =arraylist[i][0];
                iflisttext = str(arraylist[i][0])[0:20]
                lbid ='singleChoice';
                choiceid= 'singlch'+str(i)
                choices = ''
                defaultAnswer = arraylist[i][int(default_answ)-1];
                defaultIndex = 0;
                for cindex in range(0,len(colmnindex)-1):
                    if(len(arraylist[i][int(colmnindex[cindex]-1)])!=0):
                        if((arraylist[i][int(colmnindex[cindex]-1)] != 'None') and (arraylist[i][int(colmnindex[cindex]-1)] != 'nan')):
                            if(arraylist[i][int(colmnindex[cindex]-1)] == defaultAnswer):
                                defaultIndex = cindex
                            choices += '<div class="p-2 rounded radio-form"><div class="form-radio"> <input class="form-radio-input" type="radio" name="'+choiceid+'" id="'+choiceid+str((cindex+1))+'" data-type="choice" value="'+arraylist[i][int(colmnindex[cindex]-1)]+'"  onclick="inputonChange(this)" onkeypress="inputonkeypress(this)" > <label class="form-radio-label" for="'+choiceid+str((cindex+1))+'""> '+arraylist[i][int(colmnindex[cindex]-1)]+'</label> </div></div>'
                if((int(i)+1) == lengtharray):
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@singleChoice'+lbid+'" data-column="field'+idval+'"  data-fieldValue="'+arraylist[i][int(default_answ)-1]+'" default-val="'+arraylist[i][int(default_answ)-1]+'" default-val="" data-v-a="false" data-formlock="false" data-scoreobject="" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="radiochoice" data-assign="false" data-default="'+str(defaultIndex)+'" data-reason="">'+choices+'</div><div class="supportlbl"><input type="hidden" class="singlehidden" value=""><span>Description</span></div></p><p class="controls"><button class="nextButton" type="submit">Submit</button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-check-circle"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>' 
                else:
                    htmltxt = '<fieldset id="'+idval+'" data-type="'+lbid+'" data-fieldName="@singleChoice'+lbid+'"  data-column="field'+idval+'"  data-fieldValue="'+arraylist[i][int(default_answ)-1]+'" default-val="'+arraylist[i][int(default_answ)-1]+'" default-val="" data-v-a="false" data-formlock="false" data-scoreobject="" ><p>	<button class="removeButton btn-prev" type="button" aria-controls="'+prevval+'" prev-controls="'+prevval+'" onclick="prevbtnclick(this);"><i class="far fa-arrow-alt-circle-up"></i>Return</button></p><p><div class="lbl"><span>'+idText+'</span></div><div class="lbla_v" style="display:none;"><iframe src=""></iframe> </div><div class="radiochoice" data-assign="false" data-default="'+str(defaultIndex)+'" data-reason="">'+choices+'</div><div class="supportlbl"><input type="hidden" class="singlehidden" value=""><span>Description</span></div></p><p class="controls"><button class="nextButton btn-next" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="nextbtnclick(this);">Next</button><button class="btn btn-default btn-enter" type="button" aria-controls="'+idnextval+'" next-controls="'+idnextval+'" onclick="enterbtnclick(this);">Press <span class="f-string-em" >Enter</span></button> </p></fieldset>'
                    listhtml = '<li class="'+elidval+'" ><a class="lielement" id="'+elidval+'" onclick="elemclick(this.id);" data-label= "'+lbid+'"><i class="fas fa-check-circle"></i><span class="count">'+str(int(i+1))+'. </span><span class="linktext">'+iflisttext+'</span></a> <a class="delicon"  onclick="deleteclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-trash-alt"></i></a><a class="dupicon" onclick="duplicateclick(this);" data-label= "'+idval+'" data-value= "'+elidval+'"><i class="far fa-copy"></i></a></li>'
            cursor.execute("INSERT INTO `formfields`(`fieldName`, `fieldData`, `fieldsetData`, `formid`) VALUES ('"+str(elidval)+"','"+str(listhtml)+"','"+str(htmltxt)+"','"+str(formid)+"')")
            connection.commit()
    cursor.close()   
    return id
        

@csrf_exempt
def import_answer_view(request):
    if request.method == 'POST':
        url = ''
        files = request.FILES.get("file") 
        uurl = request.POST.get("url")
        uformid = request.POST.get("formid")
        uformtype = request.POST.get("formtype")
        ugoogle = request.POST.get("google")
        ufileName = request.POST.get("fileName")
        userdbId = request.session.get('dbId')
        uextension = request.POST.get("extension")
        interviewcolumns = ["S.No .","Question","Answer Type","Answer","Key words"]
        ansfileName = 'ans_link_'+ufileName
        quizcolumns = ["S.No .","Question","Answer Type","Option 1","Option 2","Option 3","Option 4","Right Answer","Reason"]
        if(ugoogle == 'yes'):
            url_1 = uurl.replace("/edit#gid=", "/export?format=csv&gid=0").replace("/edit?usp=sharing", "/export?format=csv&gid=0")
            pddatframe = pd.read_csv(url_1, header=None)
            columnvalue = []
            arrayofrows = []
            columindex = []
            default_Answer = ''
            interviewName = str(pddatframe[5][0]).strip()
            for i in range(0,len(pddatframe.columns)):
                try:
                    columnvalue.append(str(pddatframe[i][3]).strip())
                    if('Option' in str(pddatframe[i][3]).strip()):
                        columindex.append(i)
                    elif('Right Answer' in str(pddatframe[i][3]).strip()):
                        default_Answer = i;
                except:
                    pass
            errorMessage = ''
            if(uformtype == 'interview'):
                if(all(item in columnvalue for item in interviewcolumns)):
                    errorMessage = "sucess"
                    pddatframe.to_excel(settings.MEDIA_ROOT +'/Interview_Templates/'+ansfileName, index=False, header=False)
                    file_path = 'media/Interview_Templates/'+ansfileName
                    for row in range(3,len(pddatframe)):
                        colomnlist = []
                        for i in range(0,len(pddatframe.columns)):
                            try:
                                colomnlist.append(str(pddatframe[i][row]).strip())
                            except:
                                pass
                        if(len(colomnlist) != 0):
                            arrayofrows.append(colomnlist)
                    uplaodpath = uploadAsnwerpath(uformid,file_path)
                    return HttpResponse(uplaodpath)
                else:
                    errorMessage = "not valid"
                    return HttpResponse(errorMessage)
            elif(uformtype == 'quiz'):
                if(all(item in columnvalue for item in quizcolumns)):
                    errorMessage = "sucess"
                    pddatframe.to_excel(settings.MEDIA_ROOT +'/Interview_Templates/'+ansfileName, index=False, header=False)
                    file_path = 'media/Interview_Templates/'+ansfileName
                    for row in range(3,len(pddatframe)):
                        colomnlist = []
                        for i in range(0,len(pddatframe.columns)):
                            try:
                                if(len(str(pddatframe[1][row]).strip()) != 0):
                                    if((str(pddatframe[i][row]).strip() != 'nan' or (i in columindex))):
                                        colomnlist.append(str(pddatframe[i][row]).strip())
                            except:
                                pass
                        if(len(colomnlist) != 0):
                            arrayofrows.append(colomnlist)
                    uplaodpath = uploadAsnwerpath(uformid,file_path)
                    return HttpResponse(uplaodpath)
                else:
                    errorMessage = "not valid"
                    return HttpResponse(errorMessage)
        else:
            if(uextension == 'xlsx'):
                pddatframe = pd.read_excel(files.read(), header=None)
                columnvalue = []
                arrayofrows = []
                columindex = []
                default_Answer = ''
                interviewName = str(pddatframe[5][0]).strip()
                for i in range(0,len(pddatframe.columns)):
                    try:
                        columnvalue.append(str(pddatframe[i][3]).strip())
                        if('Option' in str(pddatframe[i][3]).strip()):
                            columindex.append(i)
                        elif('Right Answer' in str(pddatframe[i][3]).strip()):
                            default_Answer = i;
                    except:
                        pass
                errorMessage = ''
                if(uformtype == 'interview'):
                    if(all(item in columnvalue for item in interviewcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(pddatframe)):
                            colomnlist = []
                            for i in range(0,len(pddatframe.columns)):
                                try:
                                    colomnlist.append(str(pddatframe[i][row]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        uplaodpath = uploadAsnwerpath(uformid,file_path)
                        return HttpResponse(uplaodpath)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                elif(uformtype == 'quiz'):
                    if(all(item in columnvalue for item in quizcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(pddatframe)):
                            colomnlist = []
                            for i in range(0,len(pddatframe.columns)):
                                try:
                                    if(len(str(pddatframe[1][row]).strip()) != 0):
                                        if((str(pddatframe[i][row]).strip() != 'nan' or (i in columindex))):
                                            colomnlist.append(str(pddatframe[i][row]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        uplaodpath = uploadAsnwerpath(uformid,file_path)
                        return HttpResponse(uplaodpath)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
            elif(uextension == 'xls'):
                wb = xlrd.open_workbook(file_contents=files.read())
                sheet = wb.sheet_by_index(0)
                columnvalue = []
                arrayofrows = []
                columindex = []
                default_Answer = ''
                interviewName = str(sheet.row(0)[5].value).strip()
                for i in range(0,sheet.ncols):
                    try:
                        if(len(str(sheet.row(3)[i].value).strip()) != 0):
                            columnvalue.append(str(sheet.row(3)[i].value).strip())
                            if('Option' in str(sheet.row(3)[i].value).strip()):
                                    columindex.append(i)
                            elif('Right Answer' in str(sheet.row(3)[i].value).strip()):
                                default_Answer = i;
                    except:
                        pass
                if(uformtype == 'interview'):
                    if(all(item in columnvalue for item in interviewcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,sheet.nrows):
                            colomnlist = []
                            for i in range(0,sheet.ncols):
                                try:
                                    if(len(str(sheet.row(row)[i].value).strip()) != 0):
                                        colomnlist.append(str(sheet.row(row)[i].value).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        uplaodpath = uploadAsnwerpath(uformid,file_path)
                        return HttpResponse(uplaodpath)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                elif(uformtype == 'quiz'):
                    if(all(item in columnvalue for item in quizcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,sheet.nrows):
                            colomnlist = []
                            for i in range(0,sheet.ncols):
                                try:
                                    if(len(str(sheet.row(row)[1].value).strip()) != 0):
                                        if(len(str(sheet.row(row)[i].value).strip()) != 0 or (i in columindex)):
                                            colomnlist.append(str(sheet.row(row)[i].value).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        uplaodpath = uploadAsnwerpath(uformid,file_path)
                        return HttpResponse(uplaodpath)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                return HttpResponse('sucess')
            elif(uextension == 'csv'):
                reader = csv.reader(codecs.iterdecode(files, 'utf-8', errors='ignore'))
                csvlist = list(reader)
                columnvalue = []
                arrayofrows = []
                columindex = []
                default_Answer = ''
                interviewName = str(csvlist[0][5]).strip()
                for i in range(0,len(csvlist[3])):
                    try:
                        if(len(str(csvlist[3][i]).strip()) != 0):
                            columnvalue.append(str(csvlist[3][i]).strip())
                            if('Option' in str(csvlist[3][i]).strip()):
                                columindex.append(i)
                            elif('Right Answer' in str(csvlist[3][i]).strip()):
                                default_Answer = i;
                    except:
                        pass
                if(uformtype == 'interview'):
                    if(all(item in columnvalue for item in interviewcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(csvlist)):
                            colomnlist = []
                            for i in range(0,len(csvlist[row])):
                                try:
                                    if(len(str(csvlist[row][i]).strip()) != 0):
                                        colomnlist.append(str(csvlist[row][i]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        uplaodpath = uploadAsnwerpath(uformid,file_path)
                        return HttpResponse(uplaodpath)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                elif(uformtype == 'quiz'):
                    if(all(item in columnvalue for item in quizcolumns)):
                        errorMessage = "sucess"
                        fs= FileSystemStorage(location= settings.MEDIA_ROOT +'/Interview_Templates/')
                        file_path=fs.save(ansfileName,files) 
                        file_path = 'media/Interview_Templates/'+ansfileName
                        for row in range(3,len(csvlist)):
                            colomnlist = []
                            for i in range(0,len(csvlist[row])):
                                try:
                                    if(len(str(csvlist[row][1]).strip()) != 0):
                                        if(len(str(csvlist[row][i]).strip()) != 0 or (i in columindex)):
                                            colomnlist.append(str(csvlist[row][i]).strip())
                                except:
                                    pass
                            if(len(colomnlist) != 0):
                                arrayofrows.append(colomnlist)
                        uplaodpath = uploadAsnwerpath(uformid,file_path)
                        return HttpResponse(uplaodpath)
                    else:
                        errorMessage = "not valid"
                        return HttpResponse(errorMessage)
                return HttpResponse('sucess')

def uploadAsnwerpath(formid, link_path):
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    cursor.execute("UPDATE `forms_templates` SET `answer_link`='"+link_path+"' WHERE `id` = '"+str(formid)+"'")
    connection.commit()
    connection.close()
    return "sucess"


class publicReview_View(View):
    return_url = None
    def get(self , request,Id,formType,tblName):
        rows = []
        columns = []
        final = []
        list1 = []
        list2 = []
        colms1 = []
        fields = []
        nomineelist = ['nominee','share','nominate','send']
        revielsts = ['feedback','reviewer','review']
        interviewlist = ['interview','link']
        interpattern = '|'.join(interviewlist)
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        cursor.callproc('getdetails1', [tblName, Id,])
        rowdata = ''
        rowdata1 = ''
        for result in cursor.stored_results():
            fields = [x[0] for x in result.description]
            rowdata1=result.fetchall()
        for prow in rowdata1:
            rows = list(prow)
        idcolum = tblName+'id'
        exclst =  ['ipAddress','SubDate','reviewStatus','sent','sentDate','commReview','videoColloboration','overGrade','sentReport','ReportLink','AutoGrade',idcolum]
        cursor.close()
        cursor1 = connection.cursor()
        cursor1.callproc('getcolmnData1', [tblName, ])
        coludata = ''
        coludata1 = ''
        for result in cursor1.stored_results():
            coludata1=result.fetchall()
        for pcol in coludata1:
            columns = list(pcol)
        cursor1.close()
        df = pd.DataFrame({'ques':columns,'ans':rows,'fields':fields})
        final = []
        interlist = df.index[df['fields'].str.contains('interv_')]
        ilist = interlist.tolist()
        reviewlist = df.index[df['fields'].str.contains('review')]
        rlist = reviewlist.tolist()
        final = []
        i = 0
        list1_as_set = set(ilist)
        iilist = list1_as_set.intersection(rlist)
        clist1 = list(iilist)
        exlist =  [i for i in ilist if i not in clist1]
        for ind in df.index:
            if (ind in exlist):
                final.append({"que":(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":(df['ans'][clist1[i]]).capitalize(),"field":df['fields'][ind]})
                i=i+1
            elif (ind not in rlist)and (not any(qnomi in df['fields'][ind].lower() for qnomi in nomineelist)):
                if(not any(qnor in df['fields'][ind].lower() for qnor in revielsts) and df['fields'][ind] not in exclst):
                    final.append({"que":(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":"No","field":df['fields'][ind]})
        common_df = df.loc[df['fields'] == 'commReview']
        final.append({"que":(common_df['ques'].values[0]).replace("_"," "),"ans":common_df['ans'].values[0],"rev":"No","field":common_df['fields'].values[0]})
        if(formType == 'quiz'):
            auto_df = df.loc[df['fields'] == 'AutoGrade']
            final.append({"que":(auto_df['ques'].values[0]).replace("_"," "),"ans":auto_df['ans'].values[0],"rev":"No","field":auto_df['fields'].values[0]})
        grade_df = df.loc[df['fields'] == 'overGrade']
        final.append({"que":(grade_df['ques'].values[0]).replace("_"," "),"ans":grade_df['ans'].values[0],"rev":"No","field":grade_df['fields'].values[0]})  
        return render(request , 'public_review.html',{'data':final})


def fuctionExtractArray(url):
    exlist = ['Answer Type','Text','Email','Video','Audio','MCQ','Dropdown','Scale','True/False']
    path = urlparse(url).path
    ext = str(os.path.splitext(path)[1]).replace('.','')
    arrayofrows = []
    if(ext == 'xlsx'):
        pddatframe = pd.read_excel('/home/techtvxs/public_html/interview/' + url, header=None)
        arrayofrows = []
        srno = 0
        for row in range(3,len(pddatframe)):
            colomnlist = []
            for i in range(0,len(pddatframe.columns)):
                try:
                    if(len(str(pddatframe[1][row]).strip()) != 0):
                        if(row == 3):
                            if((str(pddatframe[i][row]).strip() not in exlist) == True and str(pddatframe[i][row]).strip() !='nan'):
                                colomnlist.append(str(pddatframe[i][row]).strip())
                        elif(str(pddatframe[i][row]).strip() not in exlist and i<= noofcolumn):
                            if(i ==0 and row != 3):
                                colomnlist.append(str(srno+1))
                                srno = srno+ 1
                            else:
                                colomnlist.append("" if str(pddatframe[i][row]).strip() =='nan' else str(pddatframe[i][row]).strip())
                except:
                    pass
            if(len(colomnlist) != 0):
                arrayofrows.append(colomnlist)
                noofcolumn = len(arrayofrows[0])
        return arrayofrows
    elif(ext == 'xls'):
        wb = xlrd.open_workbook('/home/techtvxs/public_html/interview/' + url)
        sheet = wb.sheet_by_index(0)
        arrayofrows = []
        srno = 0
        for row in range(3,sheet.nrows):
            colomnlist = []
            for i in range(0,sheet.ncols):
                try:
                    if(len(str(sheet.row(row)[1].value).strip()) != 0):
                        if(row == 3):
                            if((str(sheet.row(row)[i].value).strip() not in exlist) == True and len((sheet.row(row)[i].value).strip()) !=0):
                                colomnlist.append(str(sheet.row(row)[i].value).strip())
                        elif(str(sheet.row(row)[i].value).strip() not in exlist and i<= noofcolumn):
                            if(i ==0 and row != 3):
                                colomnlist.append(str(srno+1))
                                srno = srno+ 1
                            else:
                                colomnlist.append("" if len(str(sheet.row(row)[i].value).strip()) == 0 else str(sheet.row(row)[i].value).strip())
                except:
                    pass
            if(len(colomnlist) != 0):
                arrayofrows.append(colomnlist)
                noofcolumn = len(arrayofrows[0])
        return arrayofrows
    elif(ext == 'csv'):
        csvfile = open('/home/techtvxs/public_html/interview/' + url, 'r', encoding='ISO-8859-1', newline='')
        reader = csv.reader(csvfile)
        csvlist = list(reader)
        arrayofrows = []
        srno = 0
        for row in range(3,len(csvlist)):
            colomnlist = []
            for i in range(0,len(csvlist[row])):
                try:
                    if(len(str(csvlist[row][1]).strip()) != 0):
                        if(row == 3):
                            if((str(csvlist[row][i]).strip() not in exlist) == True and len((str(csvlist[row][i]).strip())) !=0):
                                colomnlist.append(str(csvlist[row][i]).strip())
                        elif(str(csvlist[row][i]).strip() not in exlist and i<= noofcolumn):
                            if(i ==0 and row != 3):
                                colomnlist.append(str(srno+1))
                                srno = srno+ 1
                            else:
                                colomnlist.append("" if len(str(csvlist[row][i]).strip()) == 0 else str(csvlist[row][i]).strip())
                except:
                    pass
            if(len(colomnlist) != 0):
                arrayofrows.append(colomnlist)
                noofcolumn = len(arrayofrows[0])
        return arrayofrows
    return arrayofrows

def footer(canvas, doc,txt):
    canvas.saveState()
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleN.alignment = 1
    P = Paragraph(txt,styleN)
    w, h = P.wrap(doc.width, doc.bottomMargin)
    P.drawOn(canvas, doc.leftMargin, h)
    canvas.restoreState()

def sendpdfemail(arraanswers,studId, tblName,url,formId,intervtype,name, email, formName, subAt, formLink, reviewformlink,userid):
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    cursor = connection.cursor()
    getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid = '"+str(userid)+"';"
    cursor.execute(getsettings)
    results = cursor.fetchall()
    if(cursor.rowcount == 0):
        getsettings= "SELECT `settings_Field`, `settings_Value` FROM settings_tbl WHERE userid  = 51;"
        cursor.execute(getsettings)
        results = cursor.fetchall()
    host = ''
    port = ''
    login = ''  
    password = ''
    fromaddr = '' 
    message = ''
    if(len(results) != 0):
        host = results[0][1]
        port = results[1][1]
        login = results[2][1]
        password = results[3][1]
        fromaddr =  results[4][1]
        logurl = results[17][1]
        signtxt = results[19][1]
        footertxt = results[18][1]
        pdfName = "pdf_" + formId[-4:] + "report.pdf"
        pdflink = generatepdf(arraanswers,url,intervtype,pdfName,logurl,formName,name,subAt,footertxt,signtxt)
        srptuptmt = 'UPDATE '+tblName+' SET ReportLink = "'+pdflink+'" WHERE '+tblName+'id ='+str(studId)+';'
        cursor.execute(srptuptmt)
        connection.commit()
        cursor.close()
        rpttext = results[16][1]
        rpttext = str(rpttext)
        rpttext = rpttext.replace('@Name', name)
        rpttext = rpttext.replace('@Email', email)
        rpttext = rpttext.replace('@FormName', formName)
        rpttext = rpttext.replace('@SubmittedAt', subAt)
        rpttext = rpttext.replace('@FormLink', formLink)
        rpttext = rpttext.replace("@ReviewFormLink", reviewformlink)
        rpttext = rpttext.replace("@PDFReportLink", "https://test.techdivaa.com/"+pdflink)
        mailstatus = SendEmailAct(fromaddr,email,login,password,host,port,rpttext,'Auto Genereted Report Card')
        print(mailstatus)
        return mailstatus

def generatepdf(arraanswers, url, intervtype,pdfname,logourl,interName,name,subAt,footertxt,signtxt):
    elements = []
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleH = styles['Heading1']
    styleH.alignment = 1
    logo = logourl
    im = Image("/home/techtvxs/public_html/interview/"+logo, 2.5*inch, 1*inch)
    im.hAlign = 'LEFT'
    elements.append(im)
    elements.append(Spacer(10, 30))
    data = [['Interview Name:', interName.capitalize()],
        ['Candidate Name:', name.capitalize()],
        ['Interview Date:', subAt]]
    p = Paragraph("Model Answers", styleH)
    elements.append(p)
    t = Table(data,2*[2*inch], 3*[0.3*inch])
    t.setStyle(TableStyle([
        ('TEXTCOLOR',(0,0),(-1,-1),colors.black),
        ('INNERGRID', (-1,0), (-1,-1), 0.25, colors.white),
        ('BOX',(-1,0),(-1,-1),0.25,colors.white),
        ('INNERGRID', (-3,0), (-1,-1), 0.25, colors.black),
        ('BOX', (-3,0), (-1,-1), 0.25, colors.black),
        ]))
    elements.append(t)
    intervieType = intervtype
    arraylist = fuctionExtractArray(url)
    countrow = 3
    lenarray = len(arraylist) 
    couarray = 1
    data1 = []
    if(intervieType == 'quiz'):
        arraylist[0].insert(len(arraylist[0]), 'Candidate Answer')
        tblwidth = 0.6
    elif(intervieType == 'interview'):
        arraylist[0].insert(len(arraylist[0]), 'Candidate Answer')
        tblwidth = 1.8
    data1.append(arraylist[0])
    tcolmns = LongTable(data1,(len(arraylist[0])-1)*[tblwidth*inch], 1*[0.2*inch])
    tcolmns.setStyle(TableStyle([
        ('TEXTCOLOR',(0,0),(-1,-1),colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('INNERGRID', (0,0), (-1,-1), 0.15, colors.black),
        ('BOX',(0,0),(-1,-1),0.15,colors.black),
        ]))
    elements.append(Spacer(10, 20))
    elements.append(tcolmns)
    divval = 0
    if(intervieType == 'quiz'):
        divval = 3
    elif(intervieType == 'interview'):
        divval = 1
    elbreak = int((lenarray-3)/divval) 
    elbreak = elbreak if intervieType == 'quiz' else elbreak -1
    if(len(arraylist) != 0):
        j= 0
        datarows = []
        for c in range(0,(lenarray+elbreak)): 
            if(j<countrow):
                if(intervieType == 'quiz'):
                    try:
                        varlist = list(filter(lambda x:x["ques"]== re.sub(r'[?|$|.|!]',r'', str(arraylist[couarray][1]).strip()),arraanswers))
                        arraylist[couarray].insert(len(arraylist[couarray]),varlist[0]['ans'])
                    except:
                        arraylist[couarray].insert(len(arraylist[couarray]),'')
                elif(intervieType == 'interview' or intervieType == 'viva'):
                    try:
                        varlist = list(filter(lambda x:x["ques"]== re.sub(r'[?|$|.|!]',r'', str(arraylist[couarray][1]).strip()),arraanswers))
                        arraylist[couarray].insert(len(arraylist[couarray]),varlist[0]['ans'])
                    except:
                        arraylist[couarray].insert(len(arraylist[couarray]),'')
                datarows.append(arraylist[couarray])
                j = j +1
                couarray = couarray+1
            else:
                elbreak = elbreak +1
                s = getSampleStyleSheet()
                s = s["BodyText"]
                s.wordWrap = 'CJK'
                s.fontSize = 6
                if(len(datarows) !=0):
                    datarows2 = [[Paragraph(cell, s) for cell in row] for row in datarows]
                    trows = LongTable(datarows2,(len(datarows[0])-1)*[tblwidth*inch], countrow*[2*inch])
                    trows.setStyle(TableStyle([
                        ('TEXTCOLOR',(0,0),(-1,-1),colors.black),
                        ('FONTSIZE', (0, 0), (-1, -1), 6),
                        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                        ('INNERGRID', (0,0), (-1,-1), 0.15, colors.black),
                        ('BOX',(0,0),(-1,-1),0.15,colors.black),
                        ]))
                    elements.append(trows)
                    elements.append(PageBreak())
                    logo = logourl
                    im = Image("/home/techtvxs/public_html/interview/"+ logo, 2.5*inch, 1*inch)
                    im.hAlign = 'LEFT'
                    elements.append(im)
                    elements.append(Spacer(10, 50))
                    elements.append(tcolmns)
                    datarows =[]
                    if(intervieType=='quiz'):
                        countrow = 3
                    elif(intervieType=='interview'):
                        countrow = 1
                    j = 0
    doc = SimpleDocTemplate( os.path.join(settings.MEDIA_ROOT,'Temp'+pdfname),pagesize=A4,
                leftMargin=1.5*cm, rightMargin=1.5*cm,
                topMargin=0*cm,bottomMargin=1.5*cm)
    footertxt1 =  html2text.html2text(footertxt)
    doc.build(elements, onFirstPage= partial(footer, txt=footertxt1), onLaterPages=partial(footer, txt=footertxt1))
    output = PdfFileWriter()
    new_pdf = PdfFileReader(os.path.join(settings.MEDIA_ROOT,'Temp'+pdfname), "rb")
    for k in range(0,new_pdf.numPages-1):
        if(k == (new_pdf.numPages-2)):
            sigpdf = FPDF()
            sigpdf.add_page()
            sigpdf.set_xy(20,-80)
            sigpdf.set_font('arial','B',10)
            sigpdf.cell(100,20,"Yours Truly,",0,0,'L')
            sigpdf.set_xy(20,-70)
            sigpdf.set_font('arial','B',10)
            sigpdf.cell(100,20,signtxt,0,0,'L')
            sigpdf.output(os.path.join(settings.MEDIA_ROOT,'SignTemp'+pdfname), 'F')
            sign_pdf = PdfFileReader(os.path.join(settings.MEDIA_ROOT,'SignTemp'+pdfname), "rb")
            existing_pdf = PdfFileReader(open(os.path.join(settings.MEDIA_ROOT,"talentsumo_pdf_temp.pdf"), "rb"))
            page = existing_pdf.getPage(0)
            page.mergePage(new_pdf.getPage(k))
            page.mergePage(sign_pdf.getPage(0))
            output.addPage(page)
            os.remove(os.path.join(settings.MEDIA_ROOT,'SignTemp'+pdfname))
        else:
            existing_pdf = PdfFileReader(open(os.path.join(settings.MEDIA_ROOT,"talentsumo_pdf_temp.pdf"), "rb"))
            page = existing_pdf.getPage(0)
            page.mergePage(new_pdf.getPage(k))
            output.addPage(page)
    outputStream = open(os.path.join(settings.MEDIA_ROOT,"pdf_reports/"+pdfname), "wb+")
    output.write(outputStream)
    outputStream.close()
    os.remove(os.path.join(settings.MEDIA_ROOT,'Temp'+pdfname))
    return "media/pdf_reports/"+pdfname

@csrf_exempt
def send_report_view(request):
    if request.method == 'POST':
        uformtbName = request.POST.get("formtbName")
        uformstudId = request.POST.get("formstudId")
        uformType = request.POST.get("formType")
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor7 = connection.cursor()
        cursor7.execute("SELECT t.*, m.name as name, m.email as email, m.review as review, m.nominee as nominee FROM `forms_templates` as t INNER join `form_column_mapping` as m on t.tbl_Name= m.tblName WHERE t.tbl_Name="+'"'+uformtbName+'"')
        rows = cursor7.fetchone()
        candName = rows[26]
        canEmail = rows[27]
        candNameval = ''
        canEmailval = ''
        candNameset = False
        canEmailset = False
        canSubDate = ''
        canSubDateset = False
        if(rows[25] == 'no'):
            arraylist = []
            stuarraylist = []
            stuarraylist  = getstudanswers(uformtbName,uformstudId,uformType)
            for f in enumerate(list(stuarraylist)):
                listitems = list(f)
                listkeys = list(list(listitems)[1].keys())
                listvalues = list(list(listitems)[1].values())
                if(listkeys[0]== 'ques' or listkeys[1] == 'ans' or listkeys[3] == 'field'):
                    namelist = listvalues[0]
                    valuelist   =  listvalues[1]
                    fieldlist   =  listvalues[3]
                    if(str(fieldlist).strip() == "SubDate" and canSubDateset == False):
                        canSubDate = valuelist.strip()
                        canSubDateset = True
                    elif(str(fieldlist).strip() == "overGrade" or str(fieldlist).strip() == "AutoGrade" or str(fieldlist).strip() == "commReview"):
                        pass
                    elif(str(fieldlist).strip() == candName and candNameset == False):
                        candNameval = valuelist.strip()
                        candNameset = True
                        data = {}
                        data['ques'] = re.sub(r'[?|$|.|!]',r'', str(namelist).strip())
                        data['ans'] = re.sub(r'[?|$|.|!]',r'', str(valuelist).strip())
                        arraylist.append(data)
                    elif(str(fieldlist).strip() == canEmail and canEmailset == False):
                        canEmailval = valuelist.strip()
                        canEmailset = True
                        data = {}
                        data['ques'] = re.sub(r'[?|$|.|!]',r'', str(namelist).strip())
                        data['ans'] = re.sub(r'[?|$|.|!]',r'', str(valuelist).strip())
                        arraylist.append(data)
                    else:
                        data = {}
                        data['ques'] = re.sub(r'[?|$|.|!]',r'', str(namelist).strip())
                        data['ans'] = re.sub(r'[?|$|.|!]',r'', str(valuelist).strip())
                        arraylist.append(data)
            formLink = 'https://spellcheck.techdivaa.com/'+rows[2] 
            reviewLink = 'https://test.techdivaa.com/review/'+str(uformstudId)+'/'+uformtbName+''
            reportstatus = sendpdfemail(arraylist,str(uformstudId),uformtbName,rows[18],rows[2] ,rows[19] ,candNameval,canEmailval,rows[1],canSubDate,formLink,reviewLink,rows[8])
            if(reportstatus == "mail Sent"):
                cursor10 = connection.cursor()
                srptuptmt = 'UPDATE '+uformtbName+' SET sentReport = "Yes" WHERE '+uformtbName+'id ='+uformstudId+';'
                cursor10.execute(srptuptmt)
                connection.commit()
                cursor10.close()
            else:
                print("not sent")
        cursor7.close()
        return HttpResponse('sucess')


def getstudanswers(tblName,Id,formType):
    connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
    rows = []
    columns = []
    final = []
    list1 = []
    list2 = []
    colms1 = []
    fields = []
    nomineelist = ['nominee','share','nominate','send']
    revielsts = ['feedback','reviewer','review']
    interviewlist = ['interview','link']
    interpattern = '|'.join(interviewlist)
    cursor = connection.cursor()
    cursor.callproc('getdetails1', [tblName, Id,])
    rowdata = ''
    rowdata1 = ''
    for result in cursor.stored_results():
        fields = [x[0] for x in result.description]
        rowdata1=result.fetchall()
    for prow in rowdata1:
        rows = list(prow)
    idcolum = tblName+'id'
    exclst =  ['ipAddress','reviewStatus','sent','sentDate','commReview','overGrade','sentReport','ReportLink','AutoGrade',idcolum,'videoColloboration']
    cursor.close()
    cursor1 = connection.cursor()
    cursor1.callproc('getcolmnData1', [tblName, ])
    coludata = ''
    coludata1 = ''
    for result in cursor1.stored_results():
        coludata1=result.fetchall()
    for pcol in coludata1:
        columns = list(pcol)
    df = pd.DataFrame({'ques':columns,'ans':rows,'fields':fields})
    final = []
    interlist = df.index[df['fields'].str.contains('interv_')]
    ilist = interlist.tolist()
    rlist = []
    for i, j in enumerate(ilist):
        rlist.append(int(df[df["fields"]==df['fields'][j].replace("interv_","review_")].index.values))
    for ind in df.index:
        if(str(df['ans'][ind]).strip() == 'name' or str(df['ans'][ind]).strip() == 'email' or len(str(df['ans'][ind]).strip()) == 0):
            df['ans'][ind] = 'N/A'
        else:
            df['ans'][ind] = df['ans'][ind]
            if ind not in rlist:
                if(df['fields'][ind] not in exclst):
                    if(df['fields'][ind].startswith("lang_") == False):
                        final.append({"que":str(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":"No","field":df['fields'][ind]})
    for k,ind in enumerate(interlist):
        if(str(df['ans'][rlist[k]]).strip() == 'None'):
            df['ans'][rlist[k]] = 'N/A'
        final.append({"que":(df['ques'][ind]).replace("_"," "),"ans":df['ans'][ind],"rev":df['ans'][rlist[k]],"field":df['fields'][ind]})
    common_df = df.loc[df['fields'] == 'commReview']
    if(str(common_df['ans'].values[0]).strip() == 'None'):
        common_df['ans'].values[0] = 'N/A'
    final.append({"que":(common_df['ques'].values[0]).replace("_"," "),"ans":common_df['ans'].values[0],"rev":"No","field":common_df['fields'].values[0]})
    if(formType == 'quiz'):
        auto_df = df.loc[df['fields'] == 'AutoGrade']
        final.append({"que":(auto_df['ques'].values[0]).replace("_"," "),"ans":auto_df['ans'].values[0],"rev":"No","field":auto_df['fields'].values[0]})
    grade_df = df.loc[df['fields'] == 'overGrade']
    final.append({"que":(grade_df['ques'].values[0]).replace("_"," "),"ans":grade_df['ans'].values[0],"rev":"No","field":grade_df['fields'].values[0]})
    return final




def export_Data_view(request):
    if(request.method == 'GET'):
        data = []
        fields = []
        FormName = request.GET['uFormName']
        tblName = request.GET['utblName']
        connection = mysql.connector.connect(host='',database='techtvxs_interview',user='techtvxs_phpinterview',password='(%G3CBBg}pBg')
        cursor = connection.cursor()
        try:
            cursor.callproc('get_allData', [tblName])
            rowdata = ''
            rowdata1 = ''
            for result in cursor.stored_results():
                rowdata1=result.fetchall()
            columnName = []
            for i,r in enumerate(list(rowdata1)):
                if(i ==0):
                    listde = list(r)
                    listde.insert(0,'Sr. No.')
                    for c in listde:
                        columnName.append(c)
                else:
                    listde = list(r)
                    listde.insert(0,i)
                    myDict = {}
                    for k,colum in enumerate(columnName):
                        myDict[colum] = listde[k]
                    data.append(myDict)
        except:
            data = []
        return JsonResponse(data,safe=False)



